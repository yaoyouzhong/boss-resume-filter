import bossmaster
from filtering import (
    _calc_edu_bonus,
    _keyword_found,
    _parse_candidate_salary_range,
    check_required_condition,
    filter_candidate,
    parse_experience_years,
)
from storage import load_candidates_all, save_candidates_all
from doc_parser import _extract_salary_range
from constants import SCORE_THRESHOLD_STRONG
import contextlib
import io
import json
import os
import tempfile
import threading
from pathlib import Path
from unittest.mock import patch


def test_smart_scan_stop_during_filtering_checkpoints_completed_candidates():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    stop_event = threading.Event()
    raw_candidates = [
        {"geek_id": "done", "name": "已完成", "summary": "本科，5年 Java"},
        {"geek_id": "pending", "name": "未处理", "summary": "本科，5年 Java"},
    ]
    job_info = {
        "job_id": "job-java",
        "job_name": "Java工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }

    def filter_first_then_stop(*_args, **_kwargs):
        stop_event.set()
        return True, 75, {"skill_matches": ["Java"]}

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", side_effect=filter_first_then_stop), \
         patch.object(bossmaster, "merge_candidates_all") as mock_merge:
        try:
            bossmaster.smart_scan_candidates(
                FakePage(), job_info, max_rounds=1, stop_event=stop_event
            )
        except bossmaster.StopRequested:
            pass
        else:
            raise AssertionError("stop should interrupt filtering")

    checkpoint = mock_merge.call_args
    assert [c["geek_id"] for c in checkpoint.args[0]] == ["done"]
    assert checkpoint.kwargs["replace_keys"] == {("done", "Java工程师")}


def test_rescan_log_describes_candidates_as_pending_evaluation_not_new():
    source = Path(bossmaster.__file__).read_text(encoding="utf-8")
    assert "过滤当前岗位已打招呼候选人" in source
    assert "本轮待评估" in source
    assert "新增 {len(raw_candidates)} 个" not in source


def test_scan_summary_separates_rule_passed_and_ai_final_counts():
    message = bossmaster._format_scan_summary(
        "完成",
        total_rule_passed=13,
        total_raw=375,
        total_ai_evaluated=4,
        total_ai_downgraded=1,
        total_passed=12,
        total_greeted=0,
    )

    assert message == (
        "[完成] 筛选结果\n"
        "规则筛选：13 / 375 人通过\n"
        "AI 复核：成功 4 人\n"
        "AI 淘汰：1 人\n"
        "最终保留：12 人\n"
        "本轮已联系：0 人"
    )


def test_scan_summary_without_ai_does_not_repeat_final_count():
    message = bossmaster._format_scan_summary(
        "完成",
        total_rule_passed=12,
        total_raw=100,
        total_ai_evaluated=0,
        total_ai_downgraded=0,
        total_passed=12,
        total_greeted=3,
    )

    assert message == (
        "[完成] 筛选结果\n"
        "规则筛选：12 / 100 人通过\n"
        "AI 复核：未执行\n"
        "最终保留：12 人\n"
        "本轮已联系：3 人"
    )


def test_scan_summary_exposes_ai_failures_even_when_all_evaluations_fail():
    message = bossmaster._format_scan_summary(
        "完成",
        total_rule_passed=13,
        total_raw=100,
        total_ai_evaluated=0,
        total_ai_downgraded=0,
        total_passed=13,
        total_greeted=0,
        total_ai_failed=13,
    )

    assert message == (
        "[完成] 筛选结果\n"
        "规则筛选：13 / 100 人通过\n"
        "AI 复核：成功 0 人，失败 13 人（按规则结果保留）\n"
        "AI 淘汰：0 人\n"
        "最终保留：13 人\n"
        "本轮已联系：0 人"
    )


def test_scan_outcome_summary_uses_worst_job_status():
    status, detail = bossmaster._summarize_scan_outcomes([
        {"job_name": "岗位A", "status": "complete", "reason": "检测到页面底部"},
        {"job_name": "岗位B", "status": "partial", "reason": "达到 30 轮上限"},
    ])
    assert status == "达到轮次上限"
    assert detail == (
        "岗位B：达到 30 轮上限\n"
        "说明：尚未检测到列表底部，也未连续 5 轮无新增候选人，"
        "因此不能确认候选人已全部加载。\n"
        "建议：适当提高“滚动轮次”后重新运行。"
    )

    status, _ = bossmaster._summarize_scan_outcomes([
        {"job_name": "岗位A", "status": "partial", "reason": "达到 30 轮上限"},
        {"job_name": "岗位B", "status": "interrupted", "reason": "页面已离开推荐牛人页面"},
    ])
    assert status == "扫描中断"


def test_filter_audit_detail_reports_top_rejection_reasons_and_skips():
    detail = bossmaster._format_filter_audit_detail(
        skipped_blacklisted=2,
        skipped_already_greeted=3,
        rule_failed_reasons={
            "评分不足(40-49分)": 5,
            "经验不足（要求3年以上）": 7,
            "学历不符/不足（要求本科）": 4,
            "薪资不匹配（岗位最高25K）": 2,
        },
        ai_hard_rejected=1,
        ai_score_rejected=2,
    )

    assert detail == (
        "已屏蔽跳过：2 人\n"
        "已沟通跳过：3 人\n"
        "主要淘汰原因\n"
        "- 经验不足（要求3年以上）：7 人\n"
        "- 学历不符/不足（要求本科）：4 人\n"
        "- 薪资不匹配（岗位最高25K）：2 人\n"
        "AI 淘汰：硬条件 1 人、降分 2 人"
    )


def test_selector_health_rejects_disconnected_page_before_iframe_lookup():
    class DisconnectedPage:
        def run_js(self, *_args, **_kwargs):
            raise RuntimeError("与页面的连接已断开。")

        def eles(self, *_args, **_kwargs):
            raise AssertionError("连接验证失败后不应继续检查 iframe")

    raised = False
    try:
        bossmaster.check_selectors_health(DisconnectedPage())
    except RuntimeError as exc:
        raised = "连接已断开" in str(exc)

    assert raised is True


def test_selector_health_retries_instead_of_reporting_stale_iframe_as_failure():
    class ContextLostError(Exception):
        pass

    class RefreshingFrame:
        def attr(self, name):
            return "https://www.zhipin.com/web/frame/recommend/" if name == "src" else ""

        def eles(self, _selector):
            raise ContextLostError("页面被刷新，请等待页面加载完成")

    class FakePage:
        def __init__(self):
            self.frame = RefreshingFrame()

        def run_js(self, _script):
            return 1

        def eles(self, _selector):
            return [self.frame]

    raised = False
    try:
        bossmaster.check_selectors_health(FakePage())
    except ContextLostError:
        raised = True

    assert raised is True


def test_selector_health_propagates_disconnect_during_card_check():
    class PageDisconnectedError(Exception):
        pass

    class DisconnectingFrame:
        def attr(self, name):
            return "https://www.zhipin.com/web/frame/recommend/" if name == "src" else ""

        def eles(self, _selector):
            raise PageDisconnectedError("与页面的连接已断开。")

    class FakePage:
        def __init__(self):
            self.frame = DisconnectingFrame()

        def run_js(self, _script):
            return 1

        def eles(self, _selector):
            return [self.frame]

    raised = False
    try:
        bossmaster.check_selectors_health(FakePage())
    except PageDisconnectedError:
        raised = True

    assert raised is True


def test_selector_health_skips_cards_when_account_has_no_published_job():
    class FakeFrame:
        def attr(self, name):
            return "https://www.zhipin.com/web/frame/recommend/?jobid&status=0" if name == "src" else ""

        def eles(self, _selector):
            return []

        def run_js(self, script):
            if "slice(0, 2000)" in script:
                return "推荐\n您需要先发布职位，才能查看推荐牛人\n发布职位"
            return False

    class FakePage:
        def __init__(self):
            self.frame = FakeFrame()

        def run_js(self, _script):
            return 1

        def eles(self, _selector):
            return [self.frame]

        def ele(self, _selector, timeout=0):
            return None

    results = bossmaster.check_selectors_health(FakePage())
    card_result = next(r for r in results if r["group"] == "candidate_card")

    assert card_result["status"] == "skip"
    assert "需要先发布职位" in card_result["detail"]
    assert not [r for r in results if r["status"] in ("warn", "fail")]


def test_parse_experience_years_supports_arabic_and_chinese_numbers():
    cases = {
        "3 年经验": 3,
        "三年以上 Java 经验": 3,
        "十二年开发经验": 12,
        "两年工作经验": 2,
        "没有年限": None,
    }
    for text, expected in cases.items():
        assert parse_experience_years(text) == expected, text


def test_bossmaster_keeps_filtering_compatibility_exports():
    assert bossmaster.filter_candidate is filter_candidate
    assert bossmaster.check_required_condition is check_required_condition
    assert bossmaster.parse_experience_years is parse_experience_years


def test_bossmaster_keeps_storage_compatibility_exports():
    assert bossmaster.load_candidates_all is load_candidates_all
    assert bossmaster.save_candidates_all is save_candidates_all


def test_resolve_job_name_ignores_whitespace_in_gui_selection():
    job_rules = {
        "中高级AI工程师": {},
        "证券固收业务python分析师": {},
        "AIAgent工程师": {},
    }

    assert bossmaster._resolve_job_name("AI Agent工程师", job_rules) == "AIAgent工程师"
    assert bossmaster._resolve_job_name("ai agent工程师", job_rules) == "AIAgent工程师"


def test_job_title_match_is_tolerant_but_not_overbroad():
    assert bossmaster._job_titles_match("AI Agent 工程师", "aiagent工程师") is True
    assert bossmaster._job_titles_match("Java工程师", "高级 Java 工程师") is True
    assert bossmaster._job_titles_match("Java工程师", "Java工程师助理") is False
    assert bossmaster._job_titles_match("Java工程师", "Python工程师") is False
    assert bossmaster._job_titles_match("Java工程师", "") is None


def test_recommend_page_identity_maps_job_id_to_real_boss_job_name():
    class Target:
        def run_js(self, script):
            if script == "return location.href":
                return (
                    "https://www.zhipin.com/web/frame/recommend/"
                    "?jobid=encrypted-job-1&status=0"
                )
            assert "/wapi/zpjob/job/chatted/jobList" in script
            return "AI-AGENT开发（出差南宁）"

    assert bossmaster._read_recommend_page_identity(Target()) == {
        "job_id": "encrypted-job-1",
        "job_title": "AI-AGENT开发（出差南宁）",
        "href": (
            "https://www.zhipin.com/web/frame/recommend/"
            "?jobid=encrypted-job-1&status=0"
        ),
    }


def test_recommend_page_identity_never_treats_recommend_tabs_as_job_names():
    class Target:
        def run_js(self, script):
            if script == "return location.href":
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-1"
            return "精选"

    identity = bossmaster._read_recommend_page_identity(Target())

    assert identity["job_id"] == "job-1"
    assert identity["job_title"] == ""


def test_page_job_mismatch_requires_explicit_confirmation():
    callback_args = []

    with patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_read_recommend_page_identity", return_value={
             "job_id": "job-python",
             "job_title": "Python工程师",
             "href": "https://www.zhipin.com/web/frame/recommend/?jobid=job-python",
         }):
        confirmed = bossmaster._confirm_page_job_match(
            object(),
            "Java工程师",
            confirm_callback=lambda expected, actual: callback_args.append((expected, actual)) or False,
        )

    assert confirmed is False
    assert callback_args == [("Java工程师", "Python工程师")]


def test_run_stops_before_scanning_when_job_mismatch_is_rejected():
    class Args:
        clear = False
        keep_greeted = False
        job = "Java工程师"
        greet = False
        re_greet = False
        greet_level = "normal"
        greet_names = None
        list_candidates = False
        rounds = 30
        max_candidates = 400
        dom_only = True
        listener_first = False
        verbose = False
        ai_eval = False

    progress = []
    job_rules = {
        "Java工程师": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]}
    }

    with patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster, "load_job_config", return_value=(job_rules, None)), \
         patch.object(bossmaster, "_confirm_run_job_configs", return_value=True), \
         patch.object(bossmaster, "_confirm_page_job_match", return_value=False) as mock_match, \
         patch.object(bossmaster, "smart_scan_candidates") as mock_scan, \
         patch.object(bossmaster, "_save_progress_on_exit"):
        bossmaster.run_smart_scan(
            Args(),
            existing_page=object(),
            progress_callback=lambda percentage, description: progress.append((percentage, description)),
        )

    mock_match.assert_called_once()
    mock_scan.assert_not_called()
    assert progress[-1][0] == 100
    assert progress[-1][1].startswith("[已停止]")


def test_run_smart_scan_passes_greet_confirmation_callback_to_scan():
    class Args:
        clear = False
        keep_greeted = False
        job = "Java工程师"
        greet = True
        re_greet = False
        greet_level = "normal"
        greet_names = None
        list_candidates = False
        rounds = 30
        max_candidates = 400
        dom_only = True
        listener_first = False
        verbose = False
        ai_eval = False

    job_rules = {"Java工程师": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]}}
    greet_confirm = lambda _message: True

    with patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster, "load_job_config", return_value=(job_rules, None)), \
         patch.object(bossmaster, "_confirm_run_job_configs", return_value=True), \
         patch.object(bossmaster, "_confirm_page_job_match", return_value=True), \
         patch.object(bossmaster, "smart_scan_candidates", return_value=[]) as mock_scan, \
         patch.object(bossmaster, "export_to_excel", return_value=True), \
         patch.object(bossmaster, "_save_progress_on_exit"):
        bossmaster.run_smart_scan(
            Args(),
            existing_page=object(),
            greet_confirm_callback=greet_confirm,
        )

    assert mock_scan.call_args.kwargs["greet_confirm_callback"] is greet_confirm


def test_run_smart_scan_includes_ai_failures_in_final_summary():
    class Args:
        clear = False
        keep_greeted = False
        job = "Java工程师"
        greet = False
        re_greet = False
        greet_level = "normal"
        greet_names = None
        list_candidates = False
        rounds = 30
        max_candidates = 400
        dom_only = True
        listener_first = False
        verbose = False
        ai_eval = False

    job_rules = {"Java工程师": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]}}
    progress = []

    def fake_scan(*_args, stats=None, **_kwargs):
        stats.update({
            "raw_count": 10,
            "rule_passed_count": 3,
            "passed_count": 3,
            "greeted_count": 0,
            "ai_eval_count": 0,
            "ai_downgraded": 0,
            "ai_failed_count": 3,
            "scan_status": "complete",
            "scan_reason": "检测到页面底部",
        })
        return []

    with patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster, "load_job_config", return_value=(job_rules, None)), \
         patch.object(bossmaster, "_confirm_run_job_configs", return_value=True), \
         patch.object(bossmaster, "_confirm_page_job_match", return_value=True), \
         patch.object(bossmaster, "smart_scan_candidates", side_effect=fake_scan), \
         patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "export_to_excel", return_value=True):
        bossmaster.run_smart_scan(
            Args(),
            existing_page=object(),
            progress_callback=lambda percentage, description: progress.append((percentage, description)),
        )

    assert progress[-1][0] == 100
    assert "AI 复核：成功 0 人，失败 3 人（按规则结果保留）" in progress[-1][1]


def test_run_smart_scan_stops_remaining_jobs_after_boss_access_block():
    class Args:
        clear = False
        keep_greeted = False
        job = None
        greet = False
        re_greet = False
        greet_level = "normal"
        greet_names = None
        list_candidates = False
        rounds = 30
        max_candidates = 160
        dom_only = False
        listener_first = False
        verbose = False
        ai_eval = False

    job_rules = {
        "Java工程师": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
        "Python工程师": {"min_exp": 0, "edu": "不限", "keywords": ["Python"]},
    }
    progress = []

    def fake_scan(*_args, stats=None, **_kwargs):
        stats.update({
            "raw_count": 10,
            "rule_passed_count": 2,
            "passed_count": 2,
            "greeted_count": 0,
            "ai_eval_count": 0,
            "ai_downgraded": 0,
            "ai_failed_count": 0,
            "scan_status": "interrupted",
            "scan_reason": "BOSS 访问保护触发（API 直调第 1 页，HTTP 429）",
            "boss_access_blocked": True,
        })
        return []

    with patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster, "load_job_config", return_value=(job_rules, None)), \
         patch.object(bossmaster, "_confirm_run_job_configs", return_value=True), \
         patch.object(bossmaster, "_confirm_page_job_match", return_value=True), \
         patch.object(bossmaster, "_show_job_navigation_prompt") as mock_navigation, \
         patch.object(bossmaster, "smart_scan_candidates", side_effect=fake_scan) as mock_scan, \
         patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "export_to_excel", return_value=True):
        bossmaster.run_smart_scan(
            Args(),
            existing_page=object(),
            progress_callback=lambda percentage, description: progress.append((percentage, description)),
        )

    assert mock_scan.call_count == 1
    mock_navigation.assert_not_called()
    assert progress[-1][0] == 100
    assert progress[-1][1].startswith("[扫描中断]")


def test_run_job_config_diagnostics_ignores_info_only():
    text, has_error = bossmaster._build_run_job_config_diagnostics(
        ["数据工程师"],
        {"数据工程师": {
            "min_exp": 3,
            "keywords": [
                {"name": "Python", "weight": 2},
                {"name": "SQL", "weight": 2},
                {"name": "ETL", "weight": 1},
            ],
        }},
    )

    assert text == ""
    assert has_error is False


def test_run_job_config_errors_cannot_be_overridden():
    callback_args = []
    confirmed = bossmaster._confirm_run_job_configs(
        ["Java工程师"],
        {"Java工程师": {"min_exp": "五年", "keywords": []}},
        confirm_callback=lambda text, has_error: callback_args.append((text, has_error)) or True,
    )

    assert confirmed is False
    assert callback_args[0][1] is True
    assert "最低经验不是数字" in callback_args[0][0]


def test_run_job_config_warnings_can_be_confirmed():
    confirmed = bossmaster._confirm_run_job_configs(
        ["Java工程师"],
        {"Java工程师": {
            "min_exp": 3,
            "keywords": [{"name": "Java", "weight": 2}],
            "original_requirement": "Java开发",
        }},
        confirm_callback=lambda _text, has_error: not has_error,
    )

    assert confirmed is True


def test_run_stops_before_page_check_when_job_config_is_blocked():
    class Args:
        clear = False
        keep_greeted = False
        job = "Java工程师"
        greet = False
        re_greet = False
        greet_level = "normal"
        greet_names = None
        list_candidates = False
        rounds = 30
        max_candidates = 400
        dom_only = True
        listener_first = False
        verbose = False
        ai_eval = False

    job_rules = {"Java工程师": {"min_exp": "五年", "keywords": []}}
    progress = []

    with patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster, "load_job_config", return_value=(job_rules, None)), \
         patch.object(bossmaster, "_confirm_run_job_configs", return_value=False) as mock_config, \
         patch.object(bossmaster, "_confirm_page_job_match") as mock_page_job, \
         patch.object(bossmaster, "smart_scan_candidates") as mock_scan:
        bossmaster.run_smart_scan(
            Args(),
            existing_page=object(),
            progress_callback=lambda percentage, description: progress.append((percentage, description)),
        )

    mock_config.assert_called_once()
    mock_page_job.assert_not_called()
    mock_scan.assert_not_called()
    assert progress[-1] == (100, "[已停止] 岗位配置体检未通过，未开始扫描")


def test_parse_greet_context_from_detail_url_builds_chat_start_payload():
    url = (
        "https://www.zhipin.com/wapi/zpjob/view/geek/info/v2?"
        "encryptJid=jid123&expectId=exp456&securityId=sec789&lid=lid000"
        "&entrance=0&wayType=0&sourceType=3"
    )
    context = bossmaster._parse_greet_context_from_detail_url(url)

    assert context["detail_api"]["endpoint"] == "/wapi/zpjob/view/geek/info/v2"
    assert context["detail_api"]["encryptJid"] == "jid123"
    assert context["chat_start"] == {
        "jid": "jid123",
        "expectId": "exp456",
        "lid": "lid000",
        "securityId": "sec789",
        "greet": "",
        "customGreetingGuide": "-1",
    }


def test_context_capture_skips_auto_greet_and_pending_candidates():
    candidates = [
        {"geek_id": "strong", "name": "强推", "match_score": 80, "recommend_level": "强烈推荐"},
        {"geek_id": "recommend", "name": "推荐", "match_score": 70, "recommend_level": "推荐"},
        {"geek_id": "pending", "name": "待定", "match_score": 60, "recommend_level": "待定"},
        {
            "geek_id": "review",
            "name": "待人工",
            "match_score": 80,
            "recommend_level": "强烈推荐",
            "manual_review_required": True,
        },
    ]

    selected = bossmaster._select_greet_context_candidates(
        candidates,
        auto_greet=True,
        point_to_point_mode=False,
        greet_names_list=None,
        greet_levels_allowed=["强烈推荐", "推荐"],
        existing_greeted_ids=set(),
        raw_order_by_geek_id={
            "strong": 0, "recommend": 1, "pending": 2, "review": 3,
        },
    )

    assert {c["geek_id"] for c in selected} == {"review"}


def test_context_capture_keeps_auto_greet_candidates_beyond_run_limit():
    candidates = [
        {
            "geek_id": f"g{i}",
            "name": f"候选人{i}",
            "match_score": 80,
            "recommend_level": "强烈推荐",
        }
        for i in range(52)
    ]

    selected = bossmaster._select_greet_context_candidates(
        candidates,
        auto_greet=True,
        point_to_point_mode=False,
        greet_names_list=None,
        greet_levels_allowed=["强烈推荐"],
        existing_greeted_ids=set(),
        raw_order_by_geek_id={f"g{i}": i for i in range(52)},
    )

    assert [c["geek_id"] for c in selected] == ["g50", "g51"]


def test_bind_latest_candidate_states_keeps_new_score_and_restores_manual_state():
    fresh = [{
        "geek_id": "g1",
        "job_name": "Java",
        "match_score": 80,
        "recommend_level": "强烈推荐",
        "followup_status": "未沟通",
    }]
    persisted = [{
        "geek_id": "g1",
        "job_name": "Java",
        "match_score": 60,
        "recommend_level": "待定",
        "feedback_status": "放弃",
        "feedback_updated_at": "20260729_100000",
    }]

    rebound = bossmaster._bind_latest_candidate_states(fresh, persisted)

    assert rebound[0]["match_score"] == 80
    assert rebound[0]["recommend_level"] == "强烈推荐"
    assert rebound[0]["feedback_status"] == "放弃"
    assert bossmaster.candidate_greet_skip_reason(rebound[0]) == "人工已放弃"


def test_context_capture_priority_selects_value_then_executes_page_order():
    candidates = [
        {
            "geek_id": "old-high", "match_score": 95,
            "greet_context": {"chat_start": {"jid": "old"}},
        },
        {"geek_id": "new-low", "match_score": 60},
        {"geek_id": "new-high-late", "match_score": 90},
        {"geek_id": "new-high-early", "match_score": 90},
    ]
    page_order = {
        "old-high": 0,
        "new-low": 1,
        "new-high-late": 3,
        "new-high-early": 2,
    }

    selected = bossmaster._prioritize_greet_context_candidates(
        candidates, page_order, limit=3
    )

    # 名额先给无上下文者；三人入选后，实际执行恢复为页面顺序。
    assert [c["geek_id"] for c in selected] == [
        "new-low", "new-high-early", "new-high-late",
    ]


def test_context_refresh_overwrites_old_value_only_after_success():
    old_context = {"chat_start": {"jid": "old"}}
    new_context = {"chat_start": {"jid": "new"}}
    candidate = {
        "geek_id": "g1", "name": "候选人",
        "greet_context": old_context,
        "greet_context_updated_at": "2026-06-18T10:00:00",
    }

    output = io.StringIO()
    with patch.object(
        bossmaster,
        "_capture_greet_context_from_list_page",
        side_effect=[(None, "失败"), (new_context, "成功")],
    ), patch.object(bossmaster.time, "sleep"), contextlib.redirect_stdout(output):
        first = bossmaster.enrich_greet_contexts_for_candidates(object(), [candidate], max_count=1)
        assert first == 0
        assert candidate["greet_context"] == old_context
        assert candidate["greet_context_updated_at"] == "2026-06-18T10:00:00"
        second = bossmaster.enrich_greet_contexts_for_candidates(object(), [candidate], max_count=1)

    assert second == 1
    assert candidate["greet_context"] == new_context
    assert candidate["greet_context_updated_at"] > "2026-06-18T10:00:00"
    assert "已刷新 候选人 的打招呼上下文" in output.getvalue()


def test_context_first_capture_logs_saved():
    candidate = {"geek_id": "g-new", "name": "新候选人"}
    context = {"chat_start": {"jid": "new"}}
    output = io.StringIO()
    with patch.object(
        bossmaster, "_capture_greet_context_from_list_page",
        return_value=(context, "成功"),
    ), patch.object(bossmaster.time, "sleep"), contextlib.redirect_stdout(output):
        result = bossmaster.enrich_greet_contexts_for_candidates(
            object(), [candidate], max_count=1
        )

    assert result == 1
    assert candidate["greet_context"] == context
    assert "已保存 新候选人 的打招呼上下文" in output.getvalue()


def test_context_capture_uses_slow_per_candidate_delay_and_three_item_batches():
    assert bossmaster.GREET_CONTEXT_DELAY_CENTER == 4.0
    assert bossmaster.GREET_CONTEXT_DELAY_SPREAD == 2.0
    assert bossmaster.GREET_CONTEXT_BATCH_SIZE == 3
    assert bossmaster.GREET_CONTEXT_BATCH_PAUSE_CENTER == 8.0
    assert bossmaster.GREET_CONTEXT_BATCH_PAUSE_SPREAD == 4.0


def test_dom_scan_scroll_delay_uses_named_conservative_constants():
    source = Path(bossmaster.__file__).read_text(encoding="utf-8")
    assert "_human_delay(DOM_SCROLL_DELAY_CENTER, DOM_SCROLL_DELAY_SPREAD)" in source
    assert "random.randint(DOM_SCROLL_BATCH_MIN, DOM_SCROLL_BATCH_MAX)" in source
    assert "_human_delay(DOM_SCROLL_BATCH_PAUSE_CENTER, DOM_SCROLL_BATCH_PAUSE_SPREAD)" in source
    assert "_human_delay(0.8, 0.5)" not in source


def test_smart_scan_can_disable_context_capture():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"
        listen = object()

    job_info = {
        "job_id": "job-context-off",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "g-context-off",
        "name": "张三",
        "summary": "本科，5 年 Java",
    }]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 75, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "_select_greet_context_candidates") as mock_select, \
         patch.object(bossmaster, "enrich_greet_contexts_for_candidates") as mock_enrich, \
         patch.object(bossmaster, "merge_candidates_all"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            max_rounds=1,
            greet_context_capture=False,
        )

    mock_select.assert_not_called()
    mock_enrich.assert_not_called()


def test_smart_scan_passes_context_capture_limit_to_prioritizer():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"
        listen = object()

    job_info = {
        "job_id": "job-context-limit",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": f"g-{i}", "name": f"候选人{i}", "summary": "本科，5 年 Java"}
        for i in range(3)
    ]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 75, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "_select_greet_context_candidates", return_value=raw_candidates), \
         patch.object(bossmaster, "_prioritize_greet_context_candidates", return_value=[]) as mock_prioritize, \
         patch.object(bossmaster, "merge_candidates_all"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            max_rounds=1,
            greet_context_limit=2,
        )

    assert mock_prioritize.call_args.kwargs["limit"] == 2


def test_extract_job_salary_range_handles_numeric_and_negotiable_text():
    assert _extract_salary_range("薪资范围：12k-15k") == (12, 15)
    assert _extract_salary_range("月薪: 20K-30K") == (20, 30)
    assert _extract_salary_range("薪资面议") == (None, None)
    assert _extract_salary_range("薪资可谈") == (None, None)


def test_parse_candidate_salary_range_from_summary_first_line():
    assert _parse_candidate_salary_range("15-16K\n统招本科，5 年 Java") == (15, 16)
    assert _parse_candidate_salary_range("20-35K·15薪\n本科") == (20, 35)
    assert _parse_candidate_salary_range("18K\n本科") == (18, 18)
    assert _parse_candidate_salary_range("面议\n本科") == (None, None)


def test_keyword_matching_uses_word_boundaries_for_english_terms():
    assert _keyword_found("AI Agent and LLM platform", "AI") is True
    assert _keyword_found("email platform", "AI") is False
    assert _keyword_found("熟悉智能体和知识库", "智能体") is True


def test_education_bonus_tiers_are_stable():
    assert _calc_edu_bonus("博士学历") == 10
    assert _calc_edu_bonus("211 硕士") == 9
    assert _calc_edu_bonus("硕士") == 7
    assert _calc_edu_bonus("985 本科") == 6
    assert _calc_edu_bonus("统招本科") == 3


def test_required_conditions_support_string_or_and():
    assert check_required_condition("统招本科，5 年 Java", "统招本科")["passed"] is True
    risky_result = check_required_condition("成教本科，5 年 Java", "统招本科")
    assert risky_result["passed"] is False
    assert "非统招本科" in risky_result["reason"]

    workflow = {"type": "or", "items": ["activiti", "camunda", "flowable"]}
    assert check_required_condition("有 Camunda 项目经验", workflow)["passed"] is True
    assert check_required_condition("只有 Spring Boot 经验", workflow)["passed"] is False

    stack = {"type": "and", "items": ["Java", "MySQL", "Redis"]}
    assert check_required_condition("Java MySQL Redis", stack)["passed"] is True
    assert check_required_condition("Java MySQL", stack)["passed"] is False


def test_filter_candidate_scores_and_hard_rejections_are_stable():
    rule = {
        "min_exp": 4,
        "edu": "本科",
        "work_location": "南京",
        "salary_min": 12,
        "salary_max": 15,
        "required_conditions": ["统招本科"],
        "keywords": [
            {"name": "Java", "weight": 2},
            {"name": "Spring Cloud", "weight": 2},
            {"name": "MySQL", "weight": 1},
            {"name": "Redis", "weight": 1},
        ],
    }

    passed, score, details = filter_candidate(
        "15-16K\n南京，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is True
    assert score >= SCORE_THRESHOLD_STRONG
    assert details["skill_matched_count"] == 4

    passed, _, details = filter_candidate(
        "18-22K\n南京，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is False
    assert "薪资不匹配" in details["reason"]

    passed, _, details = filter_candidate(
        "15-16K\n上海，统招本科，6 年 Java 经验，熟悉 Spring Cloud、MySQL、Redis",
        rule,
    )
    assert passed is False
    assert "地点不符" in details["reason"]


def test_geek_card_api_payload_builds_complete_candidate_summary():
    payload = {
        "zpData": {
            "geekList": [
                {
                    "encryptGeekId": "encrypted-g-api-1",
                    "geekCard": {
                        "geekId": 123456,
                        "encGeekId": "encrypted-g-api-1",
                        "geekName": "张三",
                        "ageDesc": "32岁",
                        "geekDegree": "本科",
                        "geekWorkYear": "8年",
                        "expectLocationName": "南京",
                        "salary": "15-20K",
                        "expectPositionName": "Python 后端工程师",
                        "geekDesc": {"content": "熟悉金融数据平台"},
                        "geekEdus": [
                            {
                                "school": "南京大学",
                                "major": "计算机科学",
                                "degreeName": "本科",
                                "startDate": "2008",
                                "endDate": "2012",
                            }
                        ],
                        "geekWorks": [
                            {
                                "company": "某证券公司",
                                "positionName": "高级开发工程师",
                                "responsibility": "负责 ETL 调度、Python 数据分析和 Oracle 数据库开发",
                                "workEmphasisList": ["Python", "ETL", "Oracle"],
                                "startDate": "2018",
                                "endDate": "至今",
                            }
                        ],
                    }
                }
            ]
        }
    }

    candidates = bossmaster._extract_candidates_from_api_payload(payload)

    assert candidates == [
        {
            "geek_id": "encrypted-g-api-1",
            "name": "张三",
            "summary": candidates[0]["summary"],
            "structured": candidates[0]["structured"],
            "_api_profile": candidates[0]["_api_profile"],
        }
    ]
    summary = candidates[0]["summary"]
    structured = candidates[0]["structured"]
    api_profile = candidates[0]["_api_profile"]
    assert structured.get('exp_years') == 8
    assert structured.get('age') == 32
    assert structured.get('degree') == "本科"
    assert structured.get('city') == "南京"
    # _api_profile 结构化画像
    assert api_profile['personal_summary'] == "熟悉金融数据平台"
    assert len(api_profile['educations']) == 1
    assert api_profile['educations'][0]['school'] == "南京大学"
    assert len(api_profile['works']) == 1
    assert api_profile['works'][0]['company'] == "某证券公司"
    assert api_profile['works'][0]['skills'] == ["Python", "ETL", "Oracle"]
    assert "工作职责：负责 ETL 调度、Python 数据分析和 Oracle 数据库开发" in summary
    assert "技能标签：Python、ETL、Oracle" in summary
    assert "教育经历：南京大学 计算机科学 本科 2008 2012" in summary


def test_api_candidate_summary_participates_in_existing_filtering():
    geek_card = {
        "geekId": "g-api-2",
        "geekName": "李四",
        "ageDesc": "30岁",
        "geekDegree": "本科",
        "geekWorkYear": "6年",
        "expectLocationName": "南京",
        "salary": "14-16K",
        "geekWorks": [
            {
                "positionName": "数据开发工程师",
                "responsibility": "负责 Python 爬虫、SQL 数据处理和 Agent 工作流开发",
                "workEmphasisList": [{"name": "Python"}, {"name": "SQL"}, {"name": "Agent"}],
            }
        ],
    }
    rule = {
        "min_exp": 5,
        "edu": "本科",
        "work_location": "南京",
        "salary_min": 12,
        "salary_max": 16,
        "keywords": ["Python", "SQL", "Agent"],
    }

    summary = bossmaster._build_candidate_summary_from_geek_card(geek_card)
    passed, score, details = filter_candidate(summary, rule)

    assert passed is True
    assert score >= SCORE_THRESHOLD_STRONG
    assert details["skill_matched_count"] == 3


def test_dom_scan_uses_conservative_empty_limit_without_api_listener():
    first_dom_batch = [
        {
            "geek_id": f"g-dom-{i}",
            "name": f"候选人{i}",
            "text": f"本科，{i + 4}年 Java 开发工程师",
        }
        for i in range(15)
    ]
    dom_batches = [first_dom_batch, [], [], [], [], []]

    class FakePage:
        def __init__(self):
            self.refresh_count = 0
            self.url = "https://www.zhipin.com/web/chat/recommend"
            self.url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()
    scan_stats = {}

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None) as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=([], "")) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', side_effect=dom_batches) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            page, max_rounds=6, scan_stats=scan_stats
        )

    assert len(candidates) == 15
    assert mock_dom_extract.call_count == 6
    # 无法从当前页构造 API 分页时，尝试启动 listener；启动失败后直接 DOM，不刷新页面。
    mock_start_listener.assert_called_once()
    assert page.refresh_count == 0
    mock_consume_api.assert_not_called()
    assert scan_stats["scan_status"] == "complete"
    assert scan_stats["scan_reason"] == "连续 5 轮无新增候选人"


def test_recommend_api_pagination_builds_from_current_iframe_jobid():
    class FakeFrame:
        def run_js(self, script):
            assert script == 'return location.href'
            return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"

    pagination = bossmaster._build_recommend_api_pagination_from_page(FakeFrame())

    assert pagination["base_url"] == "https://www.zhipin.com/wapi/zpjob/rec/geek/list"
    assert pagination["page_param"] == "page"
    assert pagination["page_size"] is None
    assert pagination["query_params"]["jobId"] == "job-123"
    assert pagination["query_params"]["page"] == "1"


def test_recommend_api_pagination_rejects_untrusted_hosts_and_clamps_page_size():
    assert bossmaster._looks_like_recommend_api_url(
        "https://evil.example/wapi/zpjob/rec/geek/list?page=1"
    ) is False
    assert bossmaster._parse_api_pagination(
        "https://evil.example/wapi/zpjob/rec/geek/list?page=1"
    ) is None

    pagination = bossmaster._parse_api_pagination(
        "/wapi/zpjob/rec/geek/list?page=1&pageSize=500"
    )
    assert pagination is not None
    assert pagination["base_url"] == (
        "https://www.zhipin.com/wapi/zpjob/rec/geek/list"
    )
    assert pagination["page_size"] == 20

    invalid_size = bossmaster._parse_api_pagination(
        "https://www.zhipin.com/wapi/zpjob/rec/geek/list"
        "?page=1&pageSize=invalid"
    )
    assert invalid_size is not None
    assert invalid_size["page_size"] == 20


def test_direct_api_rejects_untrusted_base_url_before_browser_fetch():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            assert False, "untrusted URLs must not reach browser fetch"

    pagination = {
        "base_url": "https://evil.example/wapi/zpjob/rec/geek/list",
        "query_params": {},
        "page_param": "page",
        "page_size": 20,
    }
    try:
        bossmaster._fetch_api_page_result(FakePage(), pagination, 1)
        assert False, "untrusted direct API URL should be rejected"
    except bossmaster.ApiClientError as exc:
        assert exc.status == "请求地址"
        assert "受信任" in exc.reason


def test_api_enrichment_keeps_dom_candidates_only():
    class FakeFrame:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakeFrame()
    dom_batch = [
        {"geek_id": "g-dom-1", "name": "张三", "text": "本科，5年 Java"},
        {"geek_id": "g-dom-2", "name": "李四", "text": "本科，6年 Java"},
    ]
    api_page = ([
        {"geek_id": "g-dom-1", "name": "张三", "summary": "本科，5年 Java，南京，25岁", "structured": {"age": 25}},
        {"geek_id": "g-api-extra", "name": "王五", "summary": "本科，8年 Java", "structured": {"exp_years": 8}},
    ], False)

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None) as mock_start_listener, \
            patch('bossmaster._fetch_api_page_result', return_value=api_page) as mock_fetch, \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            page, max_rounds=1, extraction_mode="api"
        )

    assert [c["geek_id"] for c in candidates] == ["g-dom-1", "g-dom-2"]
    assert candidates[0]["structured"] == {"age": 25}
    assert candidates[0]["summary"] == "本科，5年 Java，南京，25岁"
    assert all(c["geek_id"] != "g-api-extra" for c in candidates)
    assert page.refresh_count == 0
    mock_fetch.assert_called_once()
    mock_start_listener.assert_called_once()
    mock_consume_api.assert_not_called()
    mock_dom_extract.assert_called_once()


def test_dom_only_scan_skips_direct_api_listener_and_refresh():
    first_dom_batch = [
        {
            "geek_id": "g-dom-only-1",
            "name": "张三",
            "text": "本科，5年 Java 开发工程师",
        }
    ]

    class FakePage:
        def __init__(self):
            self.refresh_count = 0
            self.url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._build_recommend_api_pagination_from_page') as mock_build_api, \
            patch('bossmaster._fetch_api_page_result') as mock_fetch, \
            patch('bossmaster._start_recommend_api_listener') as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=first_dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            page, max_rounds=1, use_api_extraction=False
        )

    assert [c["geek_id"] for c in candidates] == ["g-dom-only-1"]
    mock_build_api.assert_not_called()
    mock_fetch.assert_not_called()
    mock_start_listener.assert_not_called()
    mock_consume_api.assert_not_called()
    assert page.refresh_count == 0
    mock_dom_extract.assert_called_once()


def test_listener_first_scan_refreshes_to_capture_first_screen_api():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            if "document.body" in script:
                return "Java 工程师 _ 南京 15-20K"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()
    dom_batch = [
        {"geek_id": "g-listener-1", "name": "李四", "text": "本科，6年 Java"},
    ]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._build_recommend_api_pagination_from_page') as mock_build_api, \
            patch('bossmaster._fetch_api_page_result') as mock_fetch, \
            patch('bossmaster._start_recommend_api_listener', return_value=FakeListener()) as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=(
                [
                    {"geek_id": "g-listener-1", "name": "李四", "summary": "本科，6年 Java", "structured": {"exp_years": 6}},
                    {"geek_id": "g-listener-2", "name": "王五", "summary": "本科，7年 Java", "structured": {"exp_years": 7}},
                ],
                "/wapi/zpjob/rec/geek/list",
            )) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            page, max_rounds=1, extraction_mode="listener", max_candidates=1
        )

    assert [c["geek_id"] for c in candidates] == ["g-listener-1"]
    assert candidates[0]["structured"] == {"exp_years": 6}
    mock_build_api.assert_not_called()
    mock_fetch.assert_not_called()
    mock_start_listener.assert_called_once()
    assert mock_consume_api.call_count >= 1
    # listener 启动后刷新一次，捕获首屏 API 响应（结构化字段来源）
    assert page.refresh_count == 1
    mock_dom_extract.assert_called_once()


def test_api_enrichment_stops_after_three_misses_within_page_cap():
    class FakeFrame:
        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

    api_pages = [
        ([{"geek_id": "g-extra-1", "name": "外部1", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-2", "name": "外部2", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-3", "name": "外部3", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-4", "name": "外部4", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-5", "name": "外部5", "summary": "本科，5年 Java", "structured": {"exp_years": 5}}], True),
    ]
    dom_batch = [{"geek_id": "g-dom-missing", "name": "张三", "text": "本科，5年 Java"}]

    with patch('bossmaster.time.sleep') as mock_sleep, \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._fetch_api_page_result', side_effect=api_pages) as mock_fetch, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakeFrame(), max_rounds=1, extraction_mode="api", max_candidates=100
        )

    # API pages return geek_ids not in DOM → matched=0 each time.
    # New logic: 3 consecutive misses → early stop (pages 4-5 never fetched).
    assert [c["geek_id"] for c in candidates] == ["g-dom-missing"]
    assert mock_fetch.call_count == 3
    assert mock_sleep.call_count >= 2
    mock_dom_extract.assert_called_once()


def test_api_enrichment_respects_one_page_cap():
    class FakeFrame:
        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

    dom_batch = [
        {"geek_id": "g-dom-1", "name": "张三", "text": "本科，5年 Java"},
        {"geek_id": "g-dom-2", "name": "李四", "text": "本科，5年 Java"},
    ]
    api_pages = [([
        {
            "geek_id": "g-dom-1",
            "name": "张三",
            "summary": "本科，5年 Java",
            "structured": {"exp_years": 5},
        }
    ], True)]

    output = io.StringIO()
    with contextlib.redirect_stdout(output), \
            patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._fetch_api_page_result', side_effect=api_pages) as mock_fetch, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch):
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakeFrame(), max_rounds=1, extraction_mode="api", max_candidates=20
        )

    assert len(candidates) == 2
    assert mock_fetch.call_count == 1
    assert "最多 1 页" in output.getvalue()


def test_api_enrichment_zero_budget_makes_zero_direct_requests():
    class FakeFrame:
        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

    dom_batch = [
        {"geek_id": "g-dom-1", "name": "张三", "text": "本科，5年 Java"},
    ]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._find_recent_recommend_api_url') as find_api_url, \
            patch('bossmaster._build_recommend_api_pagination_from_page') as build_pagination, \
            patch('bossmaster._fetch_api_page_result') as fetch_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch):
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakeFrame(), max_rounds=1, extraction_mode="api", max_candidates=0
        )

    assert [candidate["geek_id"] for candidate in candidates] == ["g-dom-1"]
    find_api_url.assert_not_called()
    build_pagination.assert_not_called()
    fetch_api.assert_not_called()


def test_default_api_enrichment_allows_eight_pages_and_warns_when_still_hitting():
    class FakeFrame:
        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

    dom_batch = [
        {"geek_id": f"g-dom-{i}", "name": f"候选人{i}", "text": "本科，5年 Java"}
        for i in range(9)
    ]
    api_pages = [
        ([
            {
                "geek_id": f"g-dom-{i}",
                "name": f"候选人{i}",
                "summary": "本科，5年 Java",
                "structured": {"exp_years": 5},
            }
        ], True)
        for i in range(8)
    ]

    output = io.StringIO()
    with contextlib.redirect_stdout(output), \
            patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._fetch_api_page_result', side_effect=api_pages) as mock_fetch, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch):
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakeFrame(), max_rounds=1, extraction_mode="api"
        )

    assert len(candidates) == 9
    assert mock_fetch.call_count == 8
    assert "最多 8 页" in output.getvalue()
    assert "API 补全已达到 8 页上限" in output.getvalue()
    assert "仍有 1 人缺少结构化信息" in output.getvalue()


def test_scan_warns_when_round_limit_ends_with_new_candidates():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    output = io.StringIO()
    scan_stats = {}
    with contextlib.redirect_stdout(output), \
            patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=[
                {"geek_id": "g-new", "name": "新增候选人", "text": "本科，5年 Java"}
            ]):
        bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(), max_rounds=1, extraction_mode="dom", scan_stats=scan_stats
        )

    assert "已达到扫描轮次上限 1" in output.getvalue()
    assert "最后一轮仍新增 1 人" in output.getvalue()
    assert scan_stats["scan_status"] == "partial"
    assert scan_stats["scan_reason"] == "达到 1 轮上限，最后一轮新增 1 人"


def test_scan_collects_last_viewport_before_stopping_at_bottom():
    class FakeFrame:
        def run_js(self, *_args, **_kwargs):
            return None

        def ele(self, *_args, **_kwargs):
            return object()

    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    frame = FakeFrame()
    first_batch = [{"geek_id": "g-first", "name": "首屏", "text": "本科，5年 Java"}]
    last_batch = first_batch + [
        {"geek_id": "g-last", "name": "末屏", "text": "本科，6年 Java"}
    ]
    scan_stats = {}

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=frame), \
            patch('bossmaster.get_frame_scroll_info', return_value={"atBottom": True}), \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', side_effect=[first_batch, last_batch]) as mock_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(), max_rounds=10, extraction_mode="dom", scan_stats=scan_stats
        )

    assert [candidate["geek_id"] for candidate in candidates] == ["g-first", "g-last"]
    assert mock_extract.call_count == 2
    assert scan_stats["scan_status"] == "complete"
    assert scan_stats["scan_reason"] == "检测到页面底部"
    assert scan_stats["scan_rounds"] == 2
    assert scan_stats["scan_last_new_count"] == 1


def test_api_enrichment_stops_after_consecutive_misses():
    """API 兜底连续 3 页无 DOM 命中时提前停止，不浪费后续请求。"""
    class FakeFrame:
        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

    # Page 1 hits (g-dom-1 in DOM), pages 2-4 miss → stop at page 4, page 5 never fetched.
    api_pages = [
        ([{"geek_id": "g-dom-1", "name": "命中", "summary": "本科，5年", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-1", "name": "外部1", "summary": "本科，5年", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-2", "name": "外部2", "summary": "本科，5年", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-3", "name": "外部3", "summary": "本科，5年", "structured": {"exp_years": 5}}], True),
        ([{"geek_id": "g-extra-4", "name": "外部4", "summary": "本科，5年", "structured": {"exp_years": 5}}], True),
    ]
    dom_batch = [
        {"geek_id": "g-dom-1", "name": "命中", "text": "本科，5年"},
        {"geek_id": "g-dom-2", "name": "李四", "text": "本科，3年"},
    ]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=None), \
            patch('bossmaster._fetch_api_page_result', side_effect=api_pages) as mock_fetch, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch):
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakeFrame(), max_rounds=1, extraction_mode="api", max_candidates=100
        )

    # Page 1 hit, pages 2-4 missed → stopped at page 4, page 5 never fetched.
    assert mock_fetch.call_count == 4
    # DOM candidates preserved (not dropped by API enrichment).
    ids = [c["geek_id"] for c in candidates]
    assert "g-dom-1" in ids
    assert "g-dom-2" in ids


def test_default_scan_uses_dom_with_listener_enrichment():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            if "document.body" in script:
                return "Java 工程师 _ 南京 15-20K"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()
    dom_batch = [{"geek_id": "g-dom-refresh", "name": "王五", "text": "本科，7年 Java"}]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=FakeListener()) as mock_start_listener, \
            patch('bossmaster._fetch_api_page_result') as mock_fetch, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=(
                [{"geek_id": "g-dom-refresh", "name": "王五", "summary": "本科，7年 Java", "structured": {"exp_years": 7}}],
                "https://www.zhipin.com/wapi/zpjob/rec/geek/list",
            )) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(page, max_rounds=1)

    assert [c["geek_id"] for c in candidates] == ["g-dom-refresh"]
    assert candidates[0]["structured"] == {"exp_years": 7}
    # listener 启动后刷新一次，捕获首屏 API 响应（结构化字段来源）
    assert page.refresh_count == 1
    mock_start_listener.assert_called_once()
    mock_fetch.assert_not_called()
    mock_consume_api.assert_called_once()
    mock_dom_extract.assert_called_once()


def test_api_risk_status_interrupts_boss_access_without_dropping_dom_candidates():
    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                return "https://www.zhipin.com/web/frame/recommend/?jobid=job-123&status=0"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()
    dom_batch = [{"geek_id": "g-dom-risk", "name": "张三", "text": "本科，5年 Java"}]
    scan_stats = {}
    output = io.StringIO()

    with contextlib.redirect_stdout(output), \
            patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch(
                'bossmaster._fetch_api_page_result',
                side_effect=bossmaster.ApiRiskBlocked(
                    429, 1, reason="请求过于频繁", source="API 直调"
                ),
            ) as mock_fetch, \
            patch('bossmaster._start_recommend_api_listener', return_value=None) as mock_start_listener, \
            patch('bossmaster._consume_recommend_api_candidates') as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            page, max_rounds=1, extraction_mode="api", scan_stats=scan_stats
        )

    assert [c["geek_id"] for c in candidates] == ["g-dom-risk"]
    assert scan_stats["scan_status"] == "interrupted"
    assert scan_stats["boss_access_blocked"] is True
    assert scan_stats["boss_block_status"] == 429
    assert "API 直调第 1 页" in scan_stats["scan_reason"]
    assert "[访问保护] BOSS 返回 HTTP 429" in output.getvalue()
    assert "将继续规则筛选、AI 评估和保存已有结果" in output.getvalue()
    assert page.refresh_count == 0
    mock_fetch.assert_called_once()
    mock_start_listener.assert_called_once()
    mock_consume_api.assert_not_called()
    mock_dom_extract.assert_called_once()


def test_boss_response_classifier_separates_global_block_and_client_errors():
    assert bossmaster._classify_boss_response(401)[0] == "block"
    assert bossmaster._classify_boss_response(418)[0] == "block"
    assert bossmaster._classify_boss_response(430)[0] == "block"
    assert bossmaster._classify_boss_response(404)[0] == "client_error"
    assert bossmaster._classify_boss_response(422)[0] == "client_error"
    assert bossmaster._classify_boss_response(503)[0] == "block"


def test_boss_response_classifier_detects_business_and_html_protection():
    action, signal, reason = bossmaster._classify_boss_response(
        200,
        body={"code": 37, "message": "检测到异常操作，请完成安全验证"},
        content_type="application/json",
    )
    assert action == "block"
    assert signal == "业务码 37"
    assert "安全验证" in reason

    action, signal, reason = bossmaster._classify_boss_response(
        200,
        body="<html><title>登录</title></html>",
        content_type="text/html",
    )
    assert action == "block"
    assert signal == "HTML 响应"
    assert "验证页" in reason


def test_boss_response_classifier_checks_body_before_ordinary_4xx():
    action, signal, reason = bossmaster._classify_boss_response(
        400,
        body={"code": 37, "message": "检测到异常操作，请完成安全验证"},
        content_type="application/json",
    )
    assert action == "block"
    assert signal == 400
    assert "安全验证" in reason

    action, signal, reason = bossmaster._classify_boss_response(
        422,
        body={"message": "登录已过期，请重新登录"},
        content_type="application/json",
    )
    assert action == "block"
    assert signal == 422
    assert "重新登录" in reason


def test_boss_response_classifier_detects_plain_text_risk_response():
    action, signal, reason = bossmaster._classify_boss_response(
        200,
        body="操作频繁，请稍后再试",
        content_type="text/plain",
    )

    assert action == "block"
    assert signal == "业务响应"
    assert "操作频繁" in reason


def test_boss_response_classifier_handles_nested_errors_and_server_failures():
    action, signal, reason = bossmaster._classify_boss_response(
        200,
        body={"error": {"code": "rate_limited", "message": "rate limit exceeded"}},
        content_type="application/json",
    )
    assert action == "block"
    assert signal == "业务码 rate_limited"
    assert "rate limit" in reason

    for status in (500, 502, 504, 511):
        action, signal, reason = bossmaster._classify_boss_response(status)
        assert action == "client_error"
        assert signal == status
        assert reason == "服务暂时不可用"


def test_boss_response_classifier_applies_risk_table_to_business_codes():
    action, signal, reason = bossmaster._classify_boss_response(
        200,
        body={"code": 403, "message": "forbidden"},
        content_type="application/json",
    )

    assert action == "block"
    assert signal == "业务码 403"
    assert reason == "forbidden"


def test_retry_after_activates_process_guard_for_followup_boss_access():
    try:
        bossmaster._raise_for_boss_response(
            429,
            body={"message": "操作频繁"},
            retry_after="7",
            source="专项测试",
        )
        assert False, "429 should activate the process-wide access guard"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.cooldown_seconds == 7

    state = bossmaster.get_boss_access_block_state()
    assert state["blocked"] is True
    assert 0 < state["remaining_seconds"] <= 7

    try:
        bossmaster._ensure_boss_access_allowed("后续访问")
        assert False, "follow-up BOSS access should be blocked during cooldown"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.from_guard is True
        assert exc.source == "后续访问"


def test_boss_access_guard_survives_restart_and_blocks_followup_access():
    with tempfile.TemporaryDirectory() as tmpdir:
        state_path = Path(tmpdir) / "boss_guard.json"
        first = bossmaster.BossAccessGuard(state_path)
        first.block(
            bossmaster.ApiRiskBlocked(
                429,
                0,
                reason="操作频繁",
                source="专项测试",
                cooldown_seconds=30,
            )
        )

        restarted = bossmaster.BossAccessGuard(state_path)
        state = restarted.state()
        assert state["blocked"] is True
        assert 20 < state["remaining_seconds"] <= 30
        try:
            restarted.ensure_allowed("重启后的访问")
            assert False, "persisted cooldown must survive process restart"
        except bossmaster.ApiRiskBlocked as exc:
            assert exc.from_guard is True
            assert exc.source == "重启后的访问"
            assert "操作频繁" in exc.reason


def test_boss_access_guard_instances_share_the_longest_cooldown():
    with tempfile.TemporaryDirectory() as tmpdir:
        state_path = Path(tmpdir) / "boss_guard.json"
        first = bossmaster.BossAccessGuard(state_path)
        second = bossmaster.BossAccessGuard(state_path)
        first.block(
            bossmaster.ApiRiskBlocked(
                429,
                0,
                reason="第一次限制",
                source="实例一",
                cooldown_seconds=5,
            )
        )
        assert second.state()["blocked"] is True

        second.block(
            bossmaster.ApiRiskBlocked(
                429,
                0,
                reason="延长限制",
                source="实例二",
                cooldown_seconds=30,
            )
        )
        refreshed = first.state()
        assert refreshed["blocked"] is True
        assert 20 < refreshed["remaining_seconds"] <= 30
        assert refreshed["source"] == "实例二"
        assert refreshed["reason"] == "延长限制"


def test_boss_access_guard_cleans_expired_persisted_state():
    with tempfile.TemporaryDirectory() as tmpdir:
        state_path = Path(tmpdir) / "boss_guard.json"
        state_path.write_text(
            json.dumps(
                {
                    "version": bossmaster.BOSS_ACCESS_GUARD_STATE_VERSION,
                    "blocked_until_epoch": 1,
                    "status": 429,
                    "reason": "已过期",
                    "source": "专项测试",
                    "triggered_at": "2026-01-01T00:00:00",
                }
            ),
            encoding="utf-8",
        )

        state = bossmaster.BossAccessGuard(state_path).state()

        assert state["blocked"] is False
        assert not state_path.exists()


def test_boss_access_guard_corrupt_state_enters_protective_cooldown():
    with tempfile.TemporaryDirectory() as tmpdir:
        state_path = Path(tmpdir) / "boss_guard.json"
        state_path.write_text("{broken", encoding="utf-8")

        state = bossmaster.BossAccessGuard(state_path).state()

        assert state["blocked"] is True
        assert state["status"] == "状态损坏"
        assert state["source"] == "本地状态恢复"
        assert "状态文件损坏" in state["reason"]
        assert state["persistence_error"]
        persisted = json.loads(state_path.read_text(encoding="utf-8"))
        assert (
            persisted["version"]
            == bossmaster.BOSS_ACCESS_GUARD_STATE_VERSION
        )


def test_boss_access_guard_read_failure_fails_closed_without_crashing():
    with tempfile.TemporaryDirectory() as tmpdir:
        guard = bossmaster.BossAccessGuard(
            Path(tmpdir) / "boss_guard.json"
        )
        with patch.object(
            guard,
            "_storage_lock",
            side_effect=OSError("permission denied"),
        ):
            state = guard.state()

        assert state["blocked"] is True
        assert state["status"] == "状态读取失败"
        assert state["source"] == "本地状态恢复"
        assert "permission denied" in state["persistence_error"]


def test_boss_control_flow_exceptions_redact_sensitive_reasons_at_source():
    client_error = bossmaster.ApiClientError(
        404,
        0,
        reason="request securityId=secret-value failed",
    )
    risk_error = bossmaster.ApiRiskBlocked(
        429,
        0,
        reason='{"expectId":"expect-secret","message":"blocked"}',
    )

    assert "secret-value" not in client_error.reason
    assert "securityId=***" in client_error.reason
    assert "expect-secret" not in risk_error.reason
    assert '"expectId":"***"' in risk_error.reason


def test_fetch_api_page_raises_client_error_for_non_risk_4xx():
    class FakePage:
        def run_js(self, _script):
            return json.dumps({
                "__bossResponse": True,
                "status": 404,
                "contentType": "application/json",
                "body": json.dumps({"message": "接口不存在"}),
            })

    pagination = {
        "base_url": "https://www.zhipin.com/wapi/zpjob/rec/geek/list",
        "query_params": {},
        "page_param": "page",
        "page_size": "20",
    }
    try:
        bossmaster._fetch_api_page_result(FakePage(), pagination, 1)
        assert False, "404 should stop the API enrichment chain"
    except bossmaster.ApiClientError as exc:
        assert exc.status == 404
        assert "接口不存在" in exc.reason


def test_listener_capture_raises_global_block_for_risk_response():
    class FakeTarget:
        def run_js(self, _script):
            return json.dumps([{
                "url": "/wapi/zpjob/rec/geek/list",
                "body": json.dumps({"message": "操作频繁，请稍后再试"}),
                "status": 429,
                "contentType": "application/json",
                "finalUrl": "https://www.zhipin.com/wapi/zpjob/rec/geek/list",
                "redirected": False,
            }])

    capture = bossmaster._ApiCapture()
    capture._page = FakeTarget()
    capture._active = True
    try:
        capture.consume()
        assert False, "listener risk response should trigger a global block"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.status == 429
        assert exc.source == "Listener"


def test_fallback_listener_stop_restores_browser_hooks_and_clears_references():
    scripts = []

    class FakeTarget:
        def run_js(self, script):
            scripts.append(script)

    target = FakeTarget()
    capture = bossmaster._ApiCapture()
    capture._page = target
    capture._active = True

    capture.stop()

    assert capture._active is False
    assert capture._page is None
    assert capture._iframe is None
    assert len(scripts) == 1
    assert "capture.requests = []" in scripts[0]
    assert "window.fetch = capture.origFetch" in scripts[0]
    assert "XMLHttpRequest.prototype.send = capture.origXhrSend" in scripts[0]


def test_listener_client_error_stops_direct_api_fallback():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def __init__(self):
            self.refresh_count = 0

        def refresh(self):
            self.refresh_count += 1

        def run_js(self, *_args, **_kwargs):
            return None

    listener_error = bossmaster.ApiClientError(
        404,
        0,
        reason="接口不存在",
        source="Listener",
    )
    with patch("bossmaster.time.sleep"), \
            patch("bossmaster._human_delay", return_value=0), \
            patch("bossmaster.get_iframe", return_value=None), \
            patch("bossmaster._read_recommend_page_identity", return_value={
                "job_id": "job-1",
                "job_title": "Java",
                "href": "https://www.zhipin.com/web/chat/recommend?jobid=job-1",
            }), \
            patch("bossmaster._start_recommend_api_listener", return_value=FakeListener()), \
            patch("bossmaster._consume_recommend_api_candidates", side_effect=listener_error), \
            patch("bossmaster._detect_captcha", return_value=(False, "")), \
            patch("bossmaster._extract_cards_batch", return_value=[{
                "geek_id": "g-listener-404",
                "name": "候选人",
                "text": "本科，5年 Java",
            }]), \
            patch("bossmaster._fetch_api_page_result") as fetch_api:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(),
            max_rounds=1,
            extraction_mode="api",
        )

    assert [candidate["geek_id"] for candidate in candidates] == ["g-listener-404"]
    fetch_api.assert_not_called()


def test_listener_identity_risk_uses_scan_block_path_before_refresh():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def __init__(self):
            self.refresh_count = 0

        def refresh(self):
            self.refresh_count += 1

    risk = bossmaster.ApiRiskBlocked(
        429,
        0,
        reason="岗位身份请求过于频繁",
        source="岗位身份读取",
    )
    scan_stats = {}
    output = io.StringIO()
    with contextlib.redirect_stdout(output), \
            patch("bossmaster.time.sleep"), \
            patch("bossmaster._human_delay", return_value=0), \
            patch("bossmaster.get_iframe", return_value=None), \
            patch("bossmaster._start_recommend_api_listener", return_value=FakeListener()), \
            patch("bossmaster._read_recommend_page_identity", side_effect=risk), \
            patch("bossmaster._extract_cards_batch") as extract_cards:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(),
            max_rounds=1,
            extraction_mode="listener",
            scan_stats=scan_stats,
        )

    assert candidates == []
    assert scan_stats["scan_status"] == "interrupted"
    assert scan_stats["boss_access_blocked"] is True
    assert scan_stats["boss_block_stage"] == "Listener 首屏"
    assert "[访问保护]" in output.getvalue()
    extract_cards.assert_not_called()


def test_detail_listener_requires_a_verified_success_response():
    detail_url = (
        "https://www.zhipin.com/wapi/zpjob/view/geek/info"
        "?jid=j1&lid=l1&securityId=s1"
    )

    class Response:
        status = 429
        headers = {
            "Content-Type": "application/json",
            "Retry-After": "9",
        }
        body = {"message": "操作频繁，请稍后再试"}
        url = detail_url

    class Packet:
        url = detail_url
        response = Response()
        is_failed = False

    class Listener:
        def wait(self, **_kwargs):
            return Packet()

    try:
        bossmaster._wait_for_detail_api_url(Listener(), timeout=0.1)
        assert False, "detail context must reject a risk response"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.status == 429
        assert exc.source == "联系信息获取"
        assert exc.cooldown_seconds == 9


def test_detail_listener_returns_url_only_after_successful_response():
    detail_url = (
        "https://www.zhipin.com/wapi/zpjob/view/geek/info"
        "?jid=j1&lid=l1&securityId=s1"
    )

    class Response:
        status = 200
        headers = {"content-type": "application/json"}
        body = {"code": 0, "zpData": {}}
        url = detail_url

    class Packet:
        url = detail_url
        response = Response()
        is_failed = False

    class Listener:
        def wait(self, **_kwargs):
            return Packet()

    assert bossmaster._wait_for_detail_api_url(Listener(), timeout=0.1) == detail_url


def test_context_greeting_risk_response_activates_guard():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return json.dumps({
                "ok": False,
                "status": 403,
                "contentType": "application/json",
                "body": json.dumps({"code": 37, "message": "操作频繁，请稍后再试"}),
            })

    context = {
        "chat_start": {
            "jid": "j1",
            "lid": "l1",
            "securityId": "secret",
        }
    }
    try:
        bossmaster.send_greeting_with_context(FakePage(), context)
        assert False, "greeting risk response should stop all BOSS access"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.status == 403
        assert exc.source == "发起沟通"
    assert bossmaster.get_boss_access_block_state()["blocked"] is True


def test_context_greeting_timeout_is_pending_verification():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            raise TimeoutError("request timeout")

    context = {
        "chat_start": {
            "jid": "j1",
            "lid": "l1",
            "securityId": "secret",
        }
    }
    success, message = bossmaster.send_greeting_with_context(FakePage(), context)
    assert success is None
    assert "待核实" in message


def test_context_greeting_pending_reason_is_redacted_before_persistence():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            raise TimeoutError("request securityId=secret-value timed out")

    context = {
        "chat_start": {
            "jid": "j1",
            "lid": "l1",
            "securityId": "secret",
        }
    }
    success, message = bossmaster.send_greeting_with_context(FakePage(), context)
    assert success is None
    assert "secret-value" not in message
    assert "securityId=***" in message


def test_context_greeting_stops_on_existing_boss_verification_page():
    class FakePage:
        url = "https://www.zhipin.com/web/user/verify?security-check=1"

        def run_js(self, *_args, **_kwargs):
            assert False, "verification pages must not issue greeting requests"

    context = {
        "chat_start": {
            "jid": "j1",
            "lid": "l1",
            "securityId": "secret",
        }
    }
    try:
        bossmaster.send_greeting_with_context(FakePage(), context)
        assert False, "existing verification page must stop context greeting"
    except bossmaster.ApiRiskBlocked as exc:
        assert exc.status == "页面跳转"
        assert exc.source == "发起沟通"


def test_list_greeting_stops_before_card_search_when_captcha_already_visible():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    with patch("bossmaster._detect_captcha", return_value=(True, "安全验证")), \
            patch("bossmaster._wait_for_captcha_resolution", return_value=True), \
            patch("bossmaster.get_iframe") as get_iframe:
        try:
            bossmaster.send_greeting_on_list_page(FakePage(), "g1")
            assert False, "visible captcha must stop before candidate card access"
        except bossmaster.ApiRiskBlocked as exc:
            assert exc.status == "安全验证"
            assert "本轮仍停止" in exc.reason

    get_iframe.assert_not_called()


def test_boss_auth_url_and_client_status_detection_are_strict():
    assert bossmaster._is_boss_auth_or_verification_url(
        "https://www.zhipin.com/web/user/verify?security-check=1"
    )
    assert not bossmaster._is_boss_auth_or_verification_url(
        "https://zhipin.com.evil.example/web/user/verify"
    )
    assert bossmaster._client_error_status_from_message(
        "上下文打招呼失败: HTTP 422 请求未成功"
    ) == 422
    assert bossmaster._client_error_status_from_message(
        "上下文打招呼失败: 业务码 422 请求未成功"
    ) == 422


def test_scan_stops_after_captcha_even_when_user_completes_it():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    scan_stats = {}
    with patch("bossmaster.time.sleep"), \
            patch("bossmaster._human_delay", return_value=0), \
            patch("bossmaster.get_iframe", return_value=None), \
            patch("bossmaster._ensure_recommend_page", return_value=True), \
            patch("bossmaster._detect_captcha", return_value=(True, "安全验证")), \
            patch("bossmaster._wait_for_captcha_resolution", return_value=True), \
            patch("bossmaster._extract_cards_batch") as extract_cards:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(),
            max_rounds=1,
            extraction_mode="dom",
            scan_stats=scan_stats,
        )

    assert candidates == []
    assert scan_stats["scan_status"] == "interrupted"
    assert scan_stats["boss_access_blocked"] is True
    assert scan_stats["boss_block_stage"] == "页面安全验证"
    extract_cards.assert_not_called()


def test_collect_captcha_diagnostic_writes_json_without_screenshot():
    class FakePage:
        def run_js(self, script):
            if script == "return location.href":
                return "https://www.zhipin.com/web/chat/recommend"
            if script == "return document.title":
                return "推荐牛人"
            if "document.body" in script:
                return "请完成安全验证"
            return ""

    with tempfile.TemporaryDirectory() as tmpdir, \
            patch.object(bossmaster, "BASE_DIR", Path(tmpdir)), \
            patch.object(bossmaster, "get_iframe", return_value=None):
        path = bossmaster._collect_captcha_diagnostic(
            FakePage(),
            detail="主页面检测到安全验证弹窗",
            stage="scan",
        )

        assert path is not None
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert payload["stage"] == "scan"
        assert payload["url"] == "https://www.zhipin.com/web/chat/recommend"
        assert "安全验证" in payload["visible_text_excerpt"]


def test_listener_mode_refreshes_and_warns_when_job_identity_changes():
    class FakeListener:
        def stop(self):
            pass

    class FakePage:
        def __init__(self):
            self.refresh_count = 0

        def run_js(self, script):
            if script == 'return location.href':
                job_id = "job-after" if self.refresh_count else "job-before"
                return f"https://www.zhipin.com/web/frame/recommend/?jobid={job_id}&status=0"
            if "document.body" in script:
                return "默认岗位 _ 南京 10-15K" if self.refresh_count else "目标岗位 _ 南京 15-20K"
            return None

        def refresh(self):
            self.refresh_count += 1

    page = FakePage()
    dom_batch = [{"geek_id": "g-stable-dom", "name": "张三", "text": "本科，5年 Java"}]

    with patch('bossmaster.time.sleep'), \
            patch('bossmaster._human_delay', return_value=0), \
            patch('bossmaster.get_iframe', return_value=None), \
            patch('bossmaster._start_recommend_api_listener', return_value=FakeListener()), \
            patch('bossmaster._fetch_api_page_result') as mock_fetch, \
            patch('bossmaster._consume_recommend_api_candidates', return_value=([], "")) as mock_consume_api, \
            patch('bossmaster._detect_captcha', return_value=(False, "")), \
            patch('bossmaster._extract_cards_batch', return_value=dom_batch) as mock_dom_extract:
        output = io.StringIO()
        with contextlib.redirect_stdout(output):
            candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
                page,
                max_rounds=1,
                extraction_mode="listener",
                blocking_notice_callback=lambda _title, _message: None,
            )

    assert [c["geek_id"] for c in candidates] == ["g-stable-dom"]
    # listener 启动后刷新一次，捕获首屏 API 响应
    assert page.refresh_count == 1
    # 刷新后岗位标识变化，应打印警告
    assert "刷新后岗位标识变化" in output.getvalue()
    mock_fetch.assert_not_called()
    # consume 被调用三次：弹窗前清空旧数据、弹窗后读取切岗新数据、
    # 滚动循环中继续消费后续数据。
    assert mock_consume_api.call_count == 3
    mock_dom_extract.assert_called_once()


def test_find_card_by_scroll_returns_to_top_after_current_position_miss():
    class FakeTarget:
        def __init__(self):
            self.scroll = 5000
            self.scripts = []

        def run_js(self, script):
            self.scripts.append(script)
            if "window.scrollTo(0, 0)" in script:
                self.scroll = 0
                return None
            if "scrollTop" in script and "return" in script:
                return self.scroll
            return None

        def ele(self, *_args, **_kwargs):
            return "card" if self.scroll == 0 else None

    target = FakeTarget()

    with patch('bossmaster.time.sleep'), patch('bossmaster._human_delay', return_value=0):
        card = bossmaster._find_card_by_scroll(target, 'css:[data-geekid="g1"]')

    assert card == "card"
    assert any("window.scrollTo(0, 0)" in script for script in target.scripts)


def test_export_to_excel_keeps_full_candidate_summary_in_detail_column():
    long_summary = (
        "15-18K\n南京，统招本科，6 年 Python 经验\n"
        + "工作职责：负责数据仓库建设、ETL 调度、SQL 优化和业务指标分析。"
        + "技能标签：Python、SQL、ETL、Oracle。"
        + "项目说明：" + "A" * 260
    )
    with tempfile.TemporaryDirectory() as tmpdir:
        output = os.path.join(tmpdir, "candidates.xlsx")
        bossmaster.export_to_excel(
            [
                {
                    "geek_id": "g-excel-1",
                    "name": "王五",
                    "summary": long_summary,
                    "job_name": "数据分析师",
                    "match_score": 80,
                    "recommend_level": "强烈推荐",
                    "greet_sent": False,
                    "next_followup_at": "20260720_090000",
                    "manual_review_required": True,
                    "risk_flags": ["学历形式待确认：疑似非统招本科"],
                    "auto_greet_blocked_reason": "学历形式待确认",
                }
            ],
            output,
        )

        from openpyxl import load_workbook

        workbook = load_workbook(output)
        assert workbook.sheetnames == ["全部候选人", "数据分析师", "统计摘要"]

        sheet = workbook["全部候选人"]
        headers = [cell.value for cell in sheet[1]]
        detail_col = headers.index("详细信息") + 1
        manual_review_col = headers.index("是否需人工确认") + 1
        risk_col = headers.index("风险提示") + 1
        next_followup_col = headers.index("下次跟进") + 1

        assert sheet.cell(row=2, column=detail_col).value == long_summary
        assert sheet.cell(row=2, column=manual_review_col).value == "是"
        assert sheet.cell(row=2, column=risk_col).value == "学历形式待确认：疑似非统招本科"
        assert sheet.cell(row=2, column=next_followup_col).value == "20260720_090000"
        assert sheet.column_dimensions[sheet.cell(row=1, column=next_followup_col).column_letter].width == 16

        job_sheet = workbook["数据分析师"]
        assert job_sheet.cell(row=2, column=1).value == 1

        summary_sheet = workbook["统计摘要"]
        summary_headers = [cell.value for cell in summary_sheet[1]]
        total_col = summary_headers.index("总人数") + 1
        avg_col = summary_headers.index("平均分") + 1
        assert summary_sheet.cell(row=2, column=total_col).value == 1
        assert summary_sheet.cell(row=2, column=avg_col).value == "80.0"


def test_export_to_excel_can_preserve_explicit_low_score_and_blacklisted_input():
    candidates = [
        {
            "geek_id": "low-score",
            "name": "低分待复核",
            "job_name": "Java",
            "match_score": 53,
            "llm_evaluated": True,
        },
        {
            "geek_id": "blacklisted",
            "name": "已屏蔽候选人",
            "job_name": "Java",
            "match_score": 70,
            "blacklisted": True,
        },
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        output = os.path.join(tmpdir, "visible-candidates.xlsx")

        assert bossmaster.export_to_excel(
            candidates,
            output,
            preserve_input=True,
        ) is True

        from openpyxl import load_workbook
        workbook = load_workbook(output)
        sheet = workbook["全部候选人"]
        headers = [cell.value for cell in sheet[1]]
        name_col = headers.index("姓名") + 1
        level_col = headers.index("推荐指数") + 1
        blacklisted_col = headers.index("是否屏蔽") + 1
        assert sheet.max_row == 3
        assert [sheet.cell(row=row, column=name_col).value for row in (2, 3)] == [
            "已屏蔽候选人",
            "低分待复核",
        ]
        assert sheet.cell(row=2, column=blacklisted_col).value == "是"
        assert sheet.cell(row=3, column=level_col).value == "未通过"


def test_auto_greet_skips_manual_review_candidates():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-risk",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {
            "min_exp": 0,
            "edu": "本科",
            "required_conditions": ["统招本科"],
            "keywords": ["Java"],
        },
    }
    raw_candidates = [{
        "geek_id": "g-risk-1",
        "name": "赵六",
        "summary": "20K\n北京，本科，5 年 Java 开发",
    }]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "send_greeting_on_list_page") as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "persist_candidate_greeted"), \
         patch.object(bossmaster.time, "sleep"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert len(result) == 1
    assert result[0]["manual_review_required"] is True
    assert result[0]["greet_sent"] is False
    mock_greet.assert_not_called()


def test_auto_greet_reuses_persisted_feedback_gate_after_rescan():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return None

    existing = [{
        "geek_id": "g-abandoned",
        "name": "已放弃候选人",
        "job_name": "Java工程师",
        "match_score": 70,
        "feedback_status": "放弃",
        "feedback_updated_at": "20260729_100000",
    }]
    raw_candidates = [{
        "geek_id": "g-abandoned",
        "name": "已放弃候选人",
        "summary": "本科，5 年 Java",
    }]
    job_info = {
        "job_id": "job-java",
        "job_name": "Java工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }

    with patch.object(bossmaster, "load_candidates_all", return_value=existing), \
         patch.object(
             bossmaster,
             "extract_candidates_by_comprehensive_analysis",
             return_value=raw_candidates,
         ), \
         patch.object(
             bossmaster,
             "filter_candidate",
             return_value=(True, 80, {"skill_matches": ["Java"]}),
         ), \
         patch.object(bossmaster, "_select_greet_context_candidates", return_value=[]), \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "send_greeting_on_list_page") as send_greeting, \
         patch.object(bossmaster.time, "sleep"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
        )

    assert result[0]["match_score"] == 80
    assert result[0]["feedback_status"] == "放弃"
    send_greeting.assert_not_called()


def test_regreet_point_and_level_filters_share_the_candidate_contact_gate():
    source = Path("bossmaster.py").read_text(encoding="utf-8")
    block = source[source.index("# 补打招呼模式：直接处理"):]
    block = block[:block.index("\n    page = None")]

    assert block.count("not candidate_greet_skip_reason(c)") >= 2


def test_smart_scan_records_filter_audit_stats():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    job_info = {
        "job_id": "job-audit",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {
            "min_exp": 3,
            "edu": "本科",
            "salary_max": 25,
            "keywords": ["Java"],
        },
    }
    raw_candidates = [
        {"geek_id": "g-pass", "name": "张三", "summary": "本科，5 年 Java"},
        {"geek_id": "g-exp", "name": "李四", "summary": "本科，1 年 Java"},
        {"geek_id": "g-score", "name": "王五", "summary": "本科，4 年 Java"},
    ]
    stats = {}

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", side_effect=[
             (True, 80, {"skill_matches": ["Java"]}),
             (False, 20, {"reason": "经验不足"}),
             (True, 52, {"skill_matches": ["Java"]}),
         ]), \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=False,
            max_rounds=1,
            stats=stats,
        )

    assert len(result) == 1
    assert stats["raw_count"] == 3
    assert stats["rule_passed_count"] == 1
    assert stats["rule_failed_count"] == 2
    assert stats["rule_failed_reasons"] == {
        "经验不足（要求3年以上）": 1,
        "评分不足(50-54分)": 1,
    }


def test_smart_scan_submits_current_results_for_existing_and_cross_job_candidates():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    existing = [
        {
            "geek_id": "same-job",
            "name": "旧同岗位记录",
            "job_name": "Java工程师",
            "match_score": 60,
            "greet_sent": False,
        },
        {
            "geek_id": "cross-job",
            "name": "跨岗位候选人",
            "job_name": "Python工程师",
            "match_score": 80,
            "greet_sent": False,
        },
    ]
    raw_candidates = [
        {"geek_id": "same-job", "name": "新同岗位记录", "summary": "本科，5年 Java"},
        {"geek_id": "cross-job", "name": "跨岗位候选人", "summary": "本科，4年 Java"},
    ]
    job_info = {
        "job_id": "job-java",
        "job_name": "Java工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }

    with patch.object(bossmaster, "load_candidates_all", return_value=existing), \
         patch.object(
             bossmaster,
             "extract_candidates_by_comprehensive_analysis",
             return_value=raw_candidates,
         ), \
         patch.object(
             bossmaster,
             "filter_candidate",
             side_effect=[
                 (True, 80, {"skill_matches": ["Java"]}),
                 (True, 75, {"skill_matches": ["Java"]}),
             ],
         ), \
         patch.object(bossmaster, "_select_greet_context_candidates", return_value=[]), \
         patch.object(bossmaster, "merge_candidates_all") as mock_merge:
        result = bossmaster.smart_scan_candidates(
            FakePage(), job_info, auto_greet=False, max_rounds=1
        )

    assert [(c["geek_id"], c["job_name"], c["match_score"]) for c in result] == [
        ("same-job", "Java工程师", 80),
        ("cross-job", "Java工程师", 75),
    ]
    submitted = mock_merge.call_args.args[0]
    assert [(c["geek_id"], c["job_name"], c["match_score"]) for c in submitted] == [
        ("same-job", "Java工程师", 80),
        ("cross-job", "Java工程师", 75),
    ]
    assert mock_merge.call_args.kwargs["replace_keys"] == {
        ("same-job", "Java工程师"),
        ("cross-job", "Java工程师"),
    }


def test_smart_scan_replaces_previous_pass_when_current_rules_reject_candidate():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    existing = [{
        "geek_id": "g-rejected",
        "name": "旧通过记录",
        "job_name": "Java工程师",
        "match_score": 80,
        "greet_sent": False,
    }]
    raw_candidates = [{
        "geek_id": "g-rejected",
        "name": "本轮未通过",
        "summary": "本科，1年 Java",
    }]
    job_info = {
        "job_id": "job-java",
        "job_name": "Java工程师",
        "rule_key": "java",
        "rule": {"min_exp": 3, "edu": "本科", "keywords": ["Java"]},
    }

    with patch.object(bossmaster, "load_candidates_all", return_value=existing), \
         patch.object(
             bossmaster,
             "extract_candidates_by_comprehensive_analysis",
             return_value=raw_candidates,
         ), \
         patch.object(
             bossmaster,
             "filter_candidate",
             return_value=(False, 25, {"reason": "经验不足"}),
         ), \
         patch.object(bossmaster, "merge_candidates_all") as mock_merge:
        result = bossmaster.smart_scan_candidates(
            FakePage(), job_info, auto_greet=False, max_rounds=1
        )

    assert result == []
    assert len(mock_merge.call_args.args[0]) == 1
    assert mock_merge.call_args.args[0][0]["qualification_status"] == "rejected"
    assert mock_merge.call_args.args[0][0]["rejection_source"] == "previously_recommended"
    assert "经验不足" in mock_merge.call_args.args[0][0]["qualification_reasons"][0]
    assert mock_merge.call_args.kwargs["replace_keys"] == {
        ("g-rejected", "Java工程师")
    }


def test_smart_scan_does_not_persist_first_scan_ordinary_rejection():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    job_info = {
        "job_id": "job-java",
        "job_name": "Java工程师",
        "rule_key": "java",
        "rule": {"min_exp": 3, "edu": "本科", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "first-reject",
        "name": "首次淘汰",
        "summary": "本科，1年 Java",
    }]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(False, 20, {"reason": "经验不足"})), \
         patch.object(bossmaster, "merge_candidates_all") as mock_merge:
        result = bossmaster.smart_scan_candidates(FakePage(), job_info, max_rounds=1)

    assert result == []
    assert mock_merge.call_args.args[0] == []


def test_auto_greet_is_disabled_when_scan_is_interrupted():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-interrupted",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "g-interrupted",
        "name": "张三",
        "summary": "本科，5 年 Java",
    }]
    stats = {}

    def fake_extract(*_args, **kwargs):
        kwargs["scan_stats"].update({
            "scan_status": "interrupted",
            "scan_reason": "页面已离开推荐牛人页面",
        })
        return raw_candidates

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", side_effect=fake_extract), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "send_greeting_on_list_page") as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all"):
        result = bossmaster.smart_scan_candidates(
            FakePage(), job_info, auto_greet=True, max_rounds=1, stats=stats
        )

    assert len(result) == 1
    assert stats["scan_status"] == "interrupted"
    mock_greet.assert_not_called()


def test_boss_access_block_keeps_ai_eval_but_skips_contact_context():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    job_info = {
        "job_id": "job-risk",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "g-risk",
        "name": "张三",
        "summary": "本科，5 年 Java",
    }]
    stats = {}

    def fake_extract(*_args, **kwargs):
        kwargs["scan_stats"].update({
            "scan_status": "interrupted",
            "scan_reason": "BOSS 访问保护触发（API 直调第 1 页，HTTP 429）",
            "boss_access_blocked": True,
            "boss_block_status": 429,
            "boss_block_stage": "API 直调第 1 页",
        })
        return raw_candidates

    def fake_ai(candidates, *_args, **_kwargs):
        result = [dict(candidate) for candidate in candidates]
        result[0]["llm_evaluated"] = True
        return result

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", side_effect=fake_extract), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch("llm_eval.evaluate_batch", side_effect=fake_ai) as mock_ai, \
         patch.object(bossmaster, "_select_greet_context_candidates") as mock_select_context, \
         patch.object(bossmaster, "merge_candidates_all"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            max_rounds=1,
            ai_eval=True,
            api_config={"model": "test"},
            api_key="test-key",
            greet_context_capture=True,
            stats=stats,
        )

    assert len(result) == 1
    assert result[0]["llm_evaluated"] is True
    mock_ai.assert_called_once()
    mock_select_context.assert_not_called()


def test_pending_snapshot_prunes_only_after_complete_scan():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

    job_info = {
        "job_id": "job-pending",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "pending",
        "name": "待定候选人",
        "summary": "本科，3年 Java",
    }]

    def run_with_status(status):
        scan_stats = {}

        def fake_extract(*_args, **kwargs):
            kwargs["scan_stats"].update({"scan_status": status, "scan_reason": status})
            return raw_candidates

        with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
             patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", side_effect=fake_extract), \
             patch.object(bossmaster, "filter_candidate", return_value=(True, 60, {"skill_matches": ["Java"]})), \
             patch.object(bossmaster, "_select_greet_context_candidates", return_value=[]), \
             patch.object(bossmaster, "merge_candidates_all") as mock_merge:
            bossmaster.smart_scan_candidates(
                FakePage(), job_info, max_rounds=1, stats=scan_stats
            )
        return mock_merge.call_args_list

    complete_calls = run_with_status("complete")
    partial_calls = run_with_status("partial")

    assert any(call.kwargs.get("prune_pending_jobs") == {"Java工程师"} for call in complete_calls)
    assert not any(call.kwargs.get("prune_pending_jobs") for call in partial_calls)


def test_auto_greet_confirmation_cancel_skips_sending_but_keeps_results():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-confirm",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [{
        "geek_id": "g-confirm",
        "name": "张三",
        "summary": "本科，5 年 Java",
    }]
    confirmations = []
    progress = []

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "send_greeting_on_list_page") as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all") as mock_merge, \
         patch.object(bossmaster.time, "sleep"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
            progress_callback=lambda pct, desc: progress.append((pct, desc)),
            greet_confirm_callback=lambda message: confirmations.append(message) or False,
        )

    assert len(result) == 1
    assert result[0]["greet_sent"] is False
    assert confirmations
    assert "本次将自动打招呼：1 人" in confirmations[0]
    assert "张三" in confirmations[0]
    mock_greet.assert_not_called()
    mock_merge.assert_called()
    assert progress[-1] == (100, "[完成] 已保存筛选结果，未执行自动打招呼")


def test_auto_greet_uses_page_order_not_score_order():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-order",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": "g-page-first", "name": "张三", "summary": "本科，3 年 Java"},
        {"geek_id": "g-page-second", "name": "李四", "summary": "本科，10 年 Java"},
    ]
    filter_results = [
        (True, 65, {"skill_matches": ["Java"]}),
        (True, 95, {"skill_matches": ["Java"]}),
    ]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", side_effect=filter_results), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "send_greeting_on_list_page", return_value=(True, "成功")) as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "persist_candidate_greeted"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert [call.args[1] for call in mock_greet.call_args_list] == [
        "g-page-first",
        "g-page-second",
    ]


def test_auto_greet_limit_triggers_notice_and_caps_greetings():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-limit",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": f"g-{i}", "name": f"候选人{i}", "summary": "本科，5 年 Java"}
        for i in range(bossmaster.AUTO_GREET_RUN_LIMIT + 5)
    ]
    notices = []

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "send_greeting_on_list_page", return_value=(True, "成功")) as mock_greet, \
         patch.object(bossmaster, "save_candidates_all"), \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "persist_candidate_greeted"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
            notice_callback=lambda title, message: notices.append((title, message)),
        )

    assert mock_greet.call_count == bossmaster.AUTO_GREET_RUN_LIMIT
    assert notices
    assert "剩余 5 人下次继续" in notices[0][1]
    assert "再次运行同一岗位扫描" in notices[0][1]


def test_auto_greet_continues_after_one_uncertain_result():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-uncertain",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": f"g-{i}", "name": f"候选人{i}", "summary": "本科，5 年 Java"}
        for i in range(3)
    ]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(
             bossmaster,
             "send_greeting_on_list_page",
             side_effect=[(None, "按钮未变化"), (True, "成功"), (True, "成功")],
         ) as mock_greet, \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "persist_candidate_greeting_pending") as mock_pending, \
         patch.object(bossmaster, "persist_candidate_greeted"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert mock_greet.call_count == 3
    mock_pending.assert_called_once()


def test_auto_greet_stops_after_two_consecutive_uncertain_results():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-uncertain-limit",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    raw_candidates = [
        {"geek_id": f"g-{i}", "name": f"候选人{i}", "summary": "本科，5 年 Java"}
        for i in range(3)
    ]

    with patch.object(bossmaster, "load_candidates_all", return_value=[]), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(
             bossmaster,
             "send_greeting_on_list_page",
             side_effect=[(None, "按钮未变化"), (None, "卡片未出现"), (True, "成功")],
         ) as mock_greet, \
         patch.object(bossmaster, "merge_candidates_all"), \
         patch.object(bossmaster, "persist_candidate_greeting_pending") as mock_pending, \
         patch.object(bossmaster, "persist_candidate_greeted"):
        bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    assert mock_greet.call_count == bossmaster.GREET_UNCERTAIN_LIMIT
    assert mock_pending.call_count == bossmaster.GREET_UNCERTAIN_LIMIT


def test_auto_greet_skips_candidate_with_existing_pending_confirmation():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/recommend"

        def run_js(self, *_args, **_kwargs):
            return None

    job_info = {
        "job_id": "job-pending-existing",
        "job_name": "Java 工程师",
        "rule_key": "java",
        "rule": {"min_exp": 0, "edu": "不限", "keywords": ["Java"]},
    }
    existing = [{
        "geek_id": "g-pending",
        "job_name": "Java工程师",
        "match_score": 80,
        "greet_sent": False,
        "greet_confirmation_pending": True,
        "greet_confirmation_reason": "按钮未变化",
        "greet_confirmation_updated_at": "20260622_100000",
    }]
    raw_candidates = [{
        "geek_id": "g-pending",
        "name": "待核实候选人",
        "summary": "本科，5 年 Java",
    }]

    with patch.object(bossmaster, "load_candidates_all", return_value=existing), \
         patch.object(bossmaster, "extract_candidates_by_comprehensive_analysis", return_value=raw_candidates), \
         patch.object(bossmaster, "filter_candidate", return_value=(True, 80, {"skill_matches": ["Java"]})), \
         patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_human_delay", return_value=0), \
         patch.object(bossmaster.time, "sleep"), \
         patch.object(bossmaster, "send_greeting_on_list_page") as mock_greet, \
         patch.object(bossmaster, "merge_candidates_all"):
        result = bossmaster.smart_scan_candidates(
            FakePage(),
            job_info,
            auto_greet=True,
            max_rounds=1,
            greet_level="normal",
        )

    mock_greet.assert_not_called()
    assert result[0]["greet_confirmation_pending"] is True


def test_limit_popup_does_not_treat_positive_remaining_count_as_exhausted():
    class FakePage:
        def run_js(self, script, *_args, **_kwargs):
            assert '"今日剩余"' in script
            assert '"今日剩余"' not in script.split('var cfg=', 1)[1].split('"upgrade"', 1)[0]
            return ""

    with patch.object(bossmaster, "get_iframe", return_value=None):
        limited, detail = bossmaster._detect_limit_popup(FakePage())

    assert limited is False
    assert detail == ""


def test_limit_popup_reports_matched_explicit_exhaustion_text():
    class FakePage:
        def run_js(self, *_args, **_kwargs):
            return json.dumps(
                {"matched": "今日沟通次数已达上限", "scope": "visible page"},
                ensure_ascii=False,
            )

    with patch.object(bossmaster, "get_iframe", return_value=None):
        limited, detail = bossmaster._detect_limit_popup(FakePage())

    assert limited is True
    assert "今日沟通次数已达上限" in detail


def test_verify_greeting_success_confirms_button_transition():
    class FakeParent:
        text = "候选人信息 继续沟通"

    class FakeCard:
        def parent(self):
            return FakeParent()

    class FakeTarget:
        def ele(self, *_args, **_kwargs):
            return FakeCard()

    success, detail = bossmaster.verify_greeting_success(
        FakeTarget(),
        "g-1",
        before_button_text="立即沟通",
        attempts=1,
        interval=0,
    )

    assert success is True
    assert "按钮已变为“继续沟通”" in detail


def test_verify_greeting_success_does_not_default_to_success_when_card_disappears():
    class FakeTarget:
        def ele(self, *_args, **_kwargs):
            return None

    success, detail = bossmaster.verify_greeting_success(
        FakeTarget(),
        "g-1",
        before_button_text="立即沟通",
        attempts=2,
        interval=0,
    )

    assert success is None
    assert "发送结果无法确认" in detail


def test_extract_stops_and_notifies_when_page_leaves_recommend_page():
    class FakePage:
        url = "https://www.zhipin.com/web/chat/index"

    notices = []

    with patch.object(bossmaster, "get_iframe", return_value=None), \
         patch.object(bossmaster, "_build_recommend_api_pagination_from_page", return_value=None), \
         patch.object(bossmaster, "_detect_captcha") as mock_captcha, \
         patch.object(bossmaster, "_extract_cards_batch") as mock_dom, \
         patch.object(bossmaster.time, "sleep"):
        result = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(),
            max_rounds=1,
            notice_callback=lambda title, message: notices.append((title, message)),
        )

    assert result == []
    assert notices
    assert notices[0][0] == "请切回推荐牛人页面"
    assert "本轮运行已停止" in notices[0][1]
    assert mock_captcha.call_count == 0
    assert mock_dom.call_count == 0


def test_ensure_recommend_page_stops_when_url_raises():
    class FakePage:
        @property
        def url(self):
            raise RuntimeError("driver lost")

        def run_js(self, *_args, **_kwargs):
            raise RuntimeError("js unavailable")

    notices = []

    ok = bossmaster._ensure_recommend_page(
        FakePage(),
        notice_callback=lambda title, message: notices.append((title, message)),
        context="扫描候选人",
    )

    assert ok is False
    assert notices
    assert "无法读取当前页面 URL" in notices[0][1]


def test_ensure_recommend_page_stops_when_url_empty():
    class FakePage:
        url = ""

        def run_js(self, *_args, **_kwargs):
            return ""

    notices = []

    ok = bossmaster._ensure_recommend_page(
        FakePage(),
        notice_callback=lambda title, message: notices.append((title, message)),
        context="扫描候选人",
    )

    assert ok is False
    assert notices
    assert "无法读取当前页面 URL" in notices[0][1]


def test_boss_verification_navigation_activates_shared_access_guard():
    class FakePage:
        url = "https://www.zhipin.com/web/user/verify?security-check=1"

    notices = []
    ok = bossmaster._ensure_recommend_page(
        FakePage(),
        notice_callback=lambda title, message: notices.append((title, message)),
        context="扫描候选人",
    )

    assert ok is False
    state = bossmaster.get_boss_access_block_state()
    assert state["blocked"] is True
    assert state["status"] == "页面跳转"
    assert state["source"] == "扫描候选人"
    assert notices


def test_dom_scan_records_verification_navigation_as_access_block():
    class FakePage:
        url = "https://www.zhipin.com/web/user/verify?security-check=1"

        def run_js(self, *_args, **_kwargs):
            return None

    scan_stats = {}
    with patch("bossmaster.time.sleep"), \
            patch("bossmaster._human_delay", return_value=0), \
            patch("bossmaster.get_iframe", return_value=None), \
            patch("bossmaster._detect_captcha") as detect_captcha, \
            patch("bossmaster._extract_cards_batch") as extract_cards:
        candidates = bossmaster.extract_candidates_by_comprehensive_analysis(
            FakePage(),
            max_rounds=1,
            extraction_mode="dom",
            scan_stats=scan_stats,
        )

    assert candidates == []
    assert scan_stats["scan_status"] == "interrupted"
    assert scan_stats["boss_access_blocked"] is True
    assert scan_stats["boss_block_stage"] == "页面跳转"
    detect_captcha.assert_not_called()
    extract_cards.assert_not_called()


def test_filter_candidate_age_boundaries_are_stable():
    rule = {
        "min_exp": 0,
        "edu": "不限",
        "max_age": 35,
        "keywords": ["Java"],
    }

    passed, _, _ = filter_candidate("35岁，Java 开发", rule)
    assert passed is True

    passed, _, details = filter_candidate("年龄：36 岁，Java 开发", rule)
    assert passed is False
    assert "年龄不符" in details["reason"]


def test_filter_candidate_rejects_upgrade_bachelor_even_with_school_mark():
    rule = {
        "min_exp": 0,
        "edu": "本科",
        "required_conditions": ["统招本科"],
        "keywords": ["Java"],
    }

    passed, _, details = filter_candidate("985 本科，专升本，5 年 Java", rule)
    assert passed is False
    assert details["qualification_status"] == "rejected"
    assert "非统招本科" in details["reason"]

    passed, _, _ = filter_candidate("全日制本科，5 年 Java", rule)
    assert passed is True


def test_save_candidates_all_deduplicates_by_geek_id_and_job_name():
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            with contextlib.redirect_stdout(io.StringIO()):
                save_candidates_all([
                    {
                        "geek_id": "g1",
                        "job_name": "Java",
                        "match_score": 70,
                        "greet_sent": True,
                        "greeting_in_progress": True,
                    },
                    {
                        "geek_id": "g1",
                        "job_name": "Java",
                        "match_score": 80,
                        "greet_sent": False,
                    },
                    {
                        "geek_id": "g1",
                        "job_name": "Python",
                        "match_score": 60,
                        "greet_sent": False,
                    },
                ])

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                saved = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert len(saved) == 2
    java = next(c for c in saved if c["job_name"] == "Java")
    python = next(c for c in saved if c["job_name"] == "Python")
    assert java["match_score"] == 80
    assert java["greet_sent"] is True
    assert "greeting_in_progress" not in java
    assert python["match_score"] == 60


def test_save_candidates_all_filters_below_55():
    """低于 55 分的候选人不应写入 candidates_all.json"""
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            with contextlib.redirect_stdout(io.StringIO()):
                save_candidates_all([
                    {"geek_id": "g1", "job_name": "Java", "match_score": 80},
                    {"geek_id": "g2", "job_name": "Java", "match_score": 55},
                    {"geek_id": "g3", "job_name": "Java", "match_score": 54},
                    {"geek_id": "g4", "job_name": "Java", "match_score": 30},
                ])

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                saved = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert len(saved) == 2
    ids = {c["geek_id"] for c in saved}
    assert ids == {"g1", "g2"}


def test_load_candidates_all_restores_from_backup_when_main_json_is_corrupt():
    old_cwd = os.getcwd()
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            os.chdir(tmpdir)
            backup_data = [{"geek_id": "g1", "job_name": "Java", "greet_sent": True}]
            with open("candidates_all.json", "w", encoding="utf-8") as f:
                f.write("{broken json")
            with open("candidates_all.json.bak", "w", encoding="utf-8") as f:
                json.dump(backup_data, f, ensure_ascii=False)

            with contextlib.redirect_stdout(io.StringIO()):
                loaded = load_candidates_all()

            with open("candidates_all.json", "r", encoding="utf-8") as f:
                restored = json.load(f)
        finally:
            os.chdir(old_cwd)

    assert loaded == backup_data
    assert restored == backup_data


def test_save_candidates_all_accepts_explicit_path():
    with tempfile.TemporaryDirectory() as tmpdir:
        target = os.path.join(tmpdir, "nested_candidates.json")
        with contextlib.redirect_stdout(io.StringIO()):
            save_candidates_all([
                {"geek_id": "g1", "job_name": "Java", "match_score": 70},
            ], target)

        with open(target, "r", encoding="utf-8") as f:
            saved = json.load(f)

    assert saved == [{"geek_id": "g1", "job_name": "Java", "match_score": 70}]


# ========== load_job_config ==========

def test_load_job_config_jobs_key_format():
    """支持 "jobs" 键格式的配置文件。"""
    config = {
        "jobs": {
            "Java工程师": {
                "min_exp": 3,
                "edu": "本科",
                "keywords": ["Java", "Spring"]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, default = load_job_config()
        assert "Java工程师" in jobs
        assert default is None
        assert jobs["Java工程师"]["min_exp"] == 3
    finally:
        os.unlink(tmp_path)


def test_load_job_config_extracts_default_rule():
    """default 规则应从 job_requirements 中提取出来单独返回。"""
    config = {
        "job_requirements": {
            "default": {"min_exp": 0, "edu": "不限", "keywords": []},
            "Python工程师": {"min_exp": 2, "keywords": ["Python"]},
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, default = load_job_config()
        assert "default" not in jobs
        assert default is not None
        assert default["edu"] == "不限"
        assert "Python工程师" in jobs
    finally:
        os.unlink(tmp_path)


def test_load_job_config_strips_spaces_from_job_name():
    """岗位名称中的空格应被移除。"""
    config = {"jobs": {"Java 工程师": {"keywords": ["Java"]}}}
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        assert "Java工程师" in jobs
        assert "Java 工程师" not in jobs
    finally:
        os.unlink(tmp_path)


def test_load_job_config_deduplicates_keywords_case_insensitive():
    """关键词应按小写去重，保留首次出现的格式。"""
    config = {
        "jobs": {
            "Dev": {
                "keywords": ["Java", "java", "JAVA", "Spring"]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        kws = jobs["Dev"]["keywords"]
        assert len(kws) == 2  # Java (case-insensitive dedup) + Spring
        assert kws[0] == "Java"  # 保留首次出现
        assert "Spring" in kws
    finally:
        os.unlink(tmp_path)


def test_load_job_config_deduplicates_dict_format_keywords():
    """dict 格式关键词也应按 name 去重。"""
    config = {
        "jobs": {
            "Dev": {
                "keywords": [
                    {"name": "Java", "weight": 2},
                    {"name": "java", "weight": 1},
                    {"name": "Spring", "weight": 1},
                ]
            }
        }
    }
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False)
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            jobs, _ = load_job_config()
        kws = jobs["Dev"]["keywords"]
        assert len(kws) == 2
        assert kws[0]["name"] == "Java"
        assert kws[0]["weight"] == 2  # 保留首次出现的 weight
    finally:
        os.unlink(tmp_path)


def test_load_job_config_missing_file_returns_default():
    """配置文件不存在时返回默认配置。"""
    with patch('bossmaster.CONFIG_PATH', '/nonexistent/path/job_config.json'):
        from bossmaster import load_job_config
        with contextlib.redirect_stdout(io.StringIO()):
            jobs, default = load_job_config()
        assert default is None
        assert "default" in jobs
        assert jobs["default"]["edu"] == "不限"


def test_load_job_config_corrupt_json_returns_default():
    """配置文件 JSON 损坏时返回默认配置。"""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        f.write("{broken json")
        tmp_path = f.name
    try:
        with patch('bossmaster.CONFIG_PATH', tmp_path):
            from bossmaster import load_job_config
            with contextlib.redirect_stdout(io.StringIO()):
                jobs, default = load_job_config()
        assert default is None
        assert "default" in jobs
    finally:
        os.unlink(tmp_path)


# ========== extract_summary_info ==========

def test_extract_summary_info_full_text():
    """完整摘要文本应提取所有字段。"""
    from bossmaster import extract_summary_info
    text = "15-20K\n30 岁，6 年经验，本科\n离职-某某科技有限公司\n南京\n熟悉 Java、Spring、MySQL、Redis"
    info = extract_summary_info(text)
    assert info['salary'] == '15-20K'
    assert info['age'] == '30'
    assert info['exp_years'] == '6'
    assert info['education'] == '本科'
    assert info['job_status'] == '离职'
    assert '某某科技' in info['company']
    assert 'Java' in info['skills']
    assert 'MySQL' in info['skills']


def test_extract_summary_info_negotiable_salary():
    """面议薪资应正确识别。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("面议\n本科，3 年经验")
    assert info['salary'] == '面议'


def test_extract_summary_info_does_not_show_age_as_experience():
    """无经验上下文的 26/27/28 年不应显示成工作年限。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("12-16K\n谭听瑞\n26年\n本科\nAI Agent Java")
    assert info['exp_years'] == ''


def test_extract_summary_info_empty_text():
    """空文本应返回全空字典。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("")
    assert all(v == '' for v in info.values())


def test_extract_summary_info_education_priority():
    """学历应取最高级别（博士 > 硕士 > 本科）。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("本科，硕士在读，博士毕业")
    assert info['education'] == '博士'


def test_extract_summary_info_status_with_company():
    """在职/离职状态和公司名提取。"""
    from bossmaster import extract_summary_info
    info = extract_summary_info("在职-阿里巴巴集团")
    assert info['job_status'] == '在职'
    assert info['company'] == '阿里巴巴集团'


# === Excel 评分拆解与简历评估的替代关系（regression: resume_adj=0 时不得回退显示 AI 调整值）===

def _export_breakdown_cell(breakdown: dict, score: int) -> str:
    with tempfile.TemporaryDirectory() as tmpdir:
        output = os.path.join(tmpdir, "c.xlsx")
        bossmaster.export_to_excel([{
            "geek_id": "g-brk",
            "name": "赵六",
            "summary": "本科\n5 年 Java",
            "job_name": "Java",
            "match_score": score,
            "recommend_level": "推荐",
            "score_breakdown": breakdown,
        }], output)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        sheet = wb["全部候选人"]
        headers = [cell.value for cell in sheet[1]]
        col = headers.index("评分拆解") + 1
        return sheet.cell(row=2, column=col).value


def _breakdown_parts_sum(line: str) -> int:
    """评分拆解行中各项（除'总分'外）的数值合计；同时识别 CJK 与拉丁（AI）标签。"""
    import re as _re
    pairs = _re.findall(r'([A-Za-z一-鿿]{2,})([-+]?\d+)', line)
    s = 0
    for label, num in pairs:
        if '总分' in label:
            continue
        s += int(num)
    return s


def test_export_excel_breakdown_resume_adj_zero_hides_ai():
    """resume_adj=0 时 Excel 评分拆解不回退显示一次评估 AI 值，合计 = 总分。"""
    breakdown = {
        "base": 25, "skill": 30, "experience": 5, "education": 5, "preferred": 0,
        "ai_adjustment": 8, "resume_adjustment": 0, "total": 65,
    }
    line = _export_breakdown_cell(breakdown, 65)
    assert "AI" not in line, f"resume_adj=0 时不应回退显示 AI 调整值：{line}"
    assert "简历" not in line, f"resume_adj=0 时不应显示简历0：{line}"
    assert _breakdown_parts_sum(line) == 65, f"拆解各项合计 != 总分 65：{line}"


def test_export_excel_breakdown_resume_adj_nonzero_shows_resume_only():
    """resume_adj≠0 时 Excel 拆解只显示简历调整值，合计 = 总分。"""
    breakdown = {
        "base": 25, "skill": 30, "experience": 5, "education": 5, "preferred": 0,
        "ai_adjustment": 8, "resume_adjustment": 5, "total": 70,
    }
    line = _export_breakdown_cell(breakdown, 70)
    assert "简历+5" in line
    assert "AI" not in line, f"有简历评估时不应显示一次评估 AI 值：{line}"
    assert _breakdown_parts_sum(line) == 70, f"拆解各项合计 != 总分 70：{line}"


def test_export_excel_breakdown_no_resume_shows_ai():
    """无简历评估时 Excel 拆解显示一次评估 AI 值，合计 = 总分。"""
    breakdown = {
        "base": 25, "skill": 30, "experience": 5, "education": 5, "preferred": 0,
        "ai_adjustment": 8, "total": 73,
    }
    line = _export_breakdown_cell(breakdown, 73)
    assert "AI+8" in line
    assert "简历" not in line
    assert _breakdown_parts_sum(line) == 73, f"拆解各项合计 != 总分 73：{line}"


# === Excel 维度列替代显示（regression: 简历评估的维度评分应替代一次评估的）===

def _export_dim_cell(c: dict) -> object:
    with tempfile.TemporaryDirectory() as tmpdir:
        output = os.path.join(tmpdir, "c.xlsx")
        bossmaster.export_to_excel([c], output)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        sheet = wb["全部候选人"]
        headers = [cell.value for cell in sheet[1]]
        col = headers.index("技能深度") + 1
        return sheet.cell(row=2, column=col).value


def test_export_excel_dim_scores_prefer_resume_over_round1():
    """Excel 维度列有简历评估时显示简历评估的值（替代一次评估的）。"""
    val = _export_dim_cell({
        'geek_id': 'g', 'name': '赵六', 'summary': '本科\n5年Java', 'job_name': 'Java',
        'match_score': 70, 'recommend_level': '推荐',
        'llm_dimension_scores': {'skill_depth': 4, 'experience_quality': 4, 'industry_fit': 4, 'growth_potential': 4},
        'resume_eval_dimension_scores': {'skill_depth': 9, 'experience_quality': 8, 'industry_fit': 7, 'growth_potential': 9},
    })
    assert val == 9, f"应显示简历评估的技能深度 9，实：{val}"


def test_export_excel_dim_scores_fallback_to_round1_without_resume():
    """无简历评估维度评分时，Excel 维度列回退显示一次评估的。"""
    val = _export_dim_cell({
        'geek_id': 'g', 'name': '赵六', 'summary': '本科\n5年Java', 'job_name': 'Java',
        'match_score': 73, 'recommend_level': '推荐',
        'llm_dimension_scores': {'skill_depth': 6, 'experience_quality': 5, 'industry_fit': 5, 'growth_potential': 5},
    })
    assert val == 6, f"应回退显示一次评估的技能深度 6，实：{val}"
